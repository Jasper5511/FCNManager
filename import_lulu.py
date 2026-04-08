# -*- coding: utf-8 -*-
"""
匯入 SN整理 - 複製.xlsx (LULU sheet) 到帳號 47 (user_id=2)
"""
import sys, io, re, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from datetime import datetime, date
from app import app
from models import db, AppUser, Client, Product, Underlying, Position

XLSX = r'C:\Users\admin\Desktop\SN整理 - 複製.xlsx'
TARGET_USER_ID = 2  # 帳號 47


def to_date(v):
    """日期轉換，處理 datetime 和多行字串"""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        # 取第一行 "2025/11/14\n4/14" → "2025/11/14"
        first_line = v.strip().split('\n')[0].strip()
        for fmt in ('%Y/%m/%d', '%Y-%m-%d'):
            try:
                return datetime.strptime(first_line, fmt).date()
            except ValueError:
                continue
    return None


def parse_ko_formula(g_val):
    """解析 KO 公式 =98%-3%*5 → (0.98, 0.03)"""
    if not g_val or not isinstance(g_val, str):
        return None, None
    m = re.match(r'=?(\d+)%-(\d+)%\*(\d+)', str(g_val))
    if m:
        start_ko = int(m.group(1)) / 100.0
        stepdown = int(m.group(2)) / 100.0
        return start_ko, stepdown
    return None, None


def parse_observation_dates(e_values, start_date):
    """解析比價日 '11/14、12/15、1/14' → ['2025-11-14', '2025-12-15', '2026-01-14']"""
    dates = []
    if not start_date:
        return dates
    start_year = start_date.year

    for e_val in e_values:
        if not e_val or not isinstance(e_val, str):
            continue
        parts = re.split(r'[、,]', e_val.strip())
        for p in parts:
            p = p.strip()
            m = re.match(r'(\d{1,2})/(\d{1,2})', p)
            if m:
                month = int(m.group(1))
                day = int(m.group(2))
                # 判斷年份：如果月份 < 起始月份，則為下一年
                if month >= start_date.month:
                    year = start_year
                else:
                    year = start_year + 1
                try:
                    dates.append(date(year, month, day).isoformat())
                except ValueError:
                    pass
    return dates


def parse_client_info(text, currency):
    """解析客戶 '姓名 金額' → ('姓名', 金額*10000)
    """
    if not text or not isinstance(text, str):
        return None, None
    text = text.strip()
    # 去掉括號備注
    clean = re.sub(r'[\(（].*?[\)）]', '', text).strip()

    # 先嘗試：中文名 + 空格/無空格 + 純數字結尾
    # 「姓名 金額」「姓名金額」
    m = re.match(r'^([\u4e00-\u9fff]{2,5})\s*(\d+\.?\d*)$', clean)
    if m:
        name = m.group(1).strip()
        amount = float(m.group(2)) * 10000  # 萬 → 實際
        return name, amount

    # 無數字或含英文備注
    m2 = re.match(r'^([\u4e00-\u9fff]{2,5})\s+(\d+\.?\d*)$', clean)
    if m2:
        return m2.group(1).strip(), float(m2.group(2)) * 10000

    # 只有中文名，無金額
    m3 = re.match(r'^([\u4e00-\u9fff]{2,5})', clean)
    if m3:
        return m3.group(1).strip(), None

    return None, None


def get_or_create_client(user_id, name):
    """取得或建立客戶，名字中間加 O"""
    masked = Client.mask_name(name)
    c = Client.query.filter_by(user_id=user_id, name=name).first()
    if not c:
        c = Client(user_id=user_id, name=name, name_masked=masked)
        db.session.add(c)
        db.session.flush()
    return c


def detect_currency(e_val):
    """從 E 欄偵測幣別"""
    if not e_val or not isinstance(e_val, str):
        return None
    val = e_val.strip().upper()
    if val in ('CNY', 'AUD', 'JPY', 'EUR', 'TWD', 'GBP'):
        return val
    return None


def import_lulu():
    wb_formula = openpyxl.load_workbook(XLSX)           # 讀公式
    wb_value = openpyxl.load_workbook(XLSX, data_only=True)  # 讀值
    ws_f = wb_formula['LULU']
    ws_v = wb_value['LULU']

    # 確認帳號 47 存在
    user = AppUser.query.get(TARGET_USER_ID)
    if not user:
        print(f'錯誤：找不到 user_id={TARGET_USER_ID}')
        return
    print(f'匯入目標帳號: {user.username} (id={user.id})')

    # 1. 找出所有商品的起始行
    product_rows = []
    for row in range(2, ws_f.max_row + 1):
        b = ws_f.cell(row, 2).value
        if b and isinstance(b, str) and ('SN' in b or 'MA' in b):
            product_rows.append(row)

    print(f'找到 {len(product_rows)} 筆商品')

    count = 0
    for idx, start_row in enumerate(product_rows):
        # 確定結束行
        end_row = product_rows[idx + 1] if idx + 1 < len(product_rows) else ws_f.max_row + 1

        # ── 基本資料 (header row) ──
        code = str(ws_f.cell(start_row, 2).value).strip()
        prod_type = ws_f.cell(start_row, 3).value or 'FCN'
        tenor = ws_v.cell(start_row, 4).value
        e_header = ws_f.cell(start_row, 5).value
        a_val = ws_f.cell(start_row, 1).value

        tickers = [
            ws_f.cell(start_row, 9).value,
            ws_f.cell(start_row, 11).value,
            ws_f.cell(start_row, 13).value,
            ws_f.cell(start_row, 15).value,
        ]

        trade_date = to_date(ws_v.cell(start_row, 16).value)
        start_date = to_date(ws_v.cell(start_row, 17).value)
        maturity_date = to_date(ws_v.cell(start_row, 18).value)
        coupon_rate = ws_v.cell(start_row, 19).value

        # 檢查帳號 47 下是否已存在
        existing = Product.query.filter_by(product_code=code, user_id=TARGET_USER_ID).first()
        if existing:
            print(f'  跳過（已存在）: {code}')
            continue


        # ── 幣別偵測 ──
        currency = detect_currency(e_header) or 'USD'

        # ── 掃描子行，分類 ──
        initial_prices = [None, None, None, None]
        ko_pct_val = None
        ko_formula = None
        ko_hits = [None, None, None, None]
        strike_pct = None
        eki_pct = None
        stepdown_text = None
        obs_date_parts = []

        for r in range(start_row, end_row):
            f_val = ws_f.cell(r, 6).value
            e_val = ws_f.cell(r, 5).value
            g_formula = ws_f.cell(r, 7).value
            g_value = ws_v.cell(r, 7).value

            # stepdown 偵測
            if e_val and isinstance(e_val, str) and 'stepdown' in e_val.lower():
                stepdown_text = e_val.strip()

            # 比價日期收集
            if e_val and isinstance(e_val, str) and '/' in e_val and 'stepdown' not in e_val.lower():
                obs_date_parts.append(e_val)

            if not f_val:
                continue
            f_str = str(f_val).strip()

            if '觀察' in f_str or '價格' in f_str:
                # 初始價格
                for i, col in enumerate([9, 11, 13, 15]):
                    v = ws_v.cell(r, col).value
                    if v and isinstance(v, (int, float)):
                        initial_prices[i] = float(v)

            elif f_str.upper() == 'KO':
                ko_pct_val = float(g_value) if g_value else None
                ko_formula = str(g_formula) if g_formula else None
                for i, col in enumerate([8, 10, 12, 14]):
                    ko_hits[i] = ws_v.cell(r, col).value

            elif f_str.upper() == 'STRIKE':
                strike_pct = float(g_value) if g_value else None

            elif f_str.upper() in ('EKI', 'KI'):
                v = g_value
                if v and isinstance(v, (int, float)) and v > 0:
                    eki_pct = float(v)

        # ── KO type & stepdown 解析 ──
        ko_type = 'fixed'
        ko_start_pct = None
        ko_stepdown_pct = None

        if stepdown_text or (e_header and isinstance(e_header, str) and 'stepdown' in e_header.lower()):
            ko_type = 'stepdown'
            sd_source = stepdown_text or e_header
            # 從 E 欄提取 stepdown 百分比
            sd_match = re.search(r'stepdown\s+(\d+)%', sd_source, re.IGNORECASE)
            sd_pct = int(sd_match.group(1)) / 100.0 if sd_match else None

            # 從 KO 公式提取起始 KO
            start, step = parse_ko_formula(ko_formula)
            if start:
                ko_start_pct = start
                ko_stepdown_pct = step
            elif sd_pct:
                ko_stepdown_pct = sd_pct
                # 無公式時從 ko_pct_val 反推（不太可靠，先存 None）
                ko_start_pct = None

        # ── 比價日 ──
        obs_dates = parse_observation_dates(obs_date_parts, start_date)

        # ── 狀態 ──
        status = 'active'
        if str(a_val).strip().upper() == 'V':
            status = 'ko_exited'

        # ── 建立 Product ──
        p = Product(
            user_id=TARGET_USER_ID,
            product_code=code,
            product_type=str(prod_type).strip(),
            tenor_months=int(tenor) if tenor else None,
            currency=currency,
            coupon_rate=float(coupon_rate) if coupon_rate else None,
            trade_date=trade_date,
            start_date=start_date,
            maturity_date=maturity_date,
            ko_pct=ko_pct_val,
            strike_pct=strike_pct,
            eki_pct=eki_pct,
            ko_type=ko_type,
            ko_start_pct=ko_start_pct,
            ko_stepdown_pct=ko_stepdown_pct,
            observation_dates=json.dumps(obs_dates) if obs_dates else None,
            ko_lockout=1,
            status=status,
        )
        db.session.add(p)
        db.session.flush()

        # ── 建立 Underlying ──
        for i, col_t in enumerate([9, 11, 13, 15]):
            ticker = tickers[i]
            if not ticker or not isinstance(ticker, str):
                continue
            init_p = initial_prices[i]
            ko_lv = round(init_p * ko_pct_val, 4) if init_p and ko_pct_val else None
            st_lv = round(init_p * strike_pct, 4) if init_p and strike_pct else None
            eki_lv = round(init_p * eki_pct, 4) if init_p and eki_pct else None

            hit_val = ko_hits[i]
            hit = str(hit_val).strip().upper() == 'V' if hit_val else False

            u = Underlying(
                product_id=p.id,
                ticker=ticker.strip().upper(),
                initial_price=init_p,
                ko_level=ko_lv,
                strike_level=st_lv,
                eki_level=eki_lv,
                ko_hit=hit,
                position_order=i + 1,
            )
            db.session.add(u)

        # ── 建立客戶 & 持倉 ──
        clients_added = []
        for r in range(start_row, end_row):
            t_val = ws_v.cell(r, 20).value  # T 欄
            if not t_val:
                continue
            name, amount = parse_client_info(str(t_val), currency)
            if not name:
                continue
            client = get_or_create_client(TARGET_USER_ID, name)
            # 避免同一商品重複加同一客戶
            dup = Position.query.filter_by(client_id=client.id, product_id=p.id).first()
            if dup:
                continue
            pos = Position(
                client_id=client.id,
                product_id=p.id,
                investment_amount=amount,
            )
            db.session.add(pos)
            clients_added.append(f'{client.name_masked}({amount})')

        count += 1
        ko_info = f'stepdown {ko_start_pct}/{ko_stepdown_pct}' if ko_type == 'stepdown' else f'fixed {ko_pct_val}'
        print(f'  {code} | {currency} | {tenor}M | KO={ko_info} | {status} | 客戶: {clients_added}')

    db.session.commit()
    print(f'\n完成！共匯入 {count} 筆商品到帳號 {user.username}')
    print(f'  商品: {Product.query.filter_by(user_id=TARGET_USER_ID).count()}')
    print(f'  客戶: {Client.query.filter_by(user_id=TARGET_USER_ID).count()}')
    print(f'  持倉: {Position.query.join(Product).filter(Product.user_id==TARGET_USER_ID).count()}')


if __name__ == '__main__':
    with app.app_context():
        import_lulu()
