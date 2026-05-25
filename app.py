import sys, io, os
os.environ['PYTHONIOENCODING'] = 'utf-8'

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, abort
from functools import wraps
from sqlalchemy.orm import joinedload
from models import db, Client, Product, Underlying, Position, PriceHistory, AppUser, ActivityLog
from config import config
from datetime import date, datetime, timedelta
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.config.from_object(config[os.environ.get('FLASK_ENV', 'default')])
if hasattr(config[os.environ.get('FLASK_ENV', 'default')], 'init_app'):
    config[os.environ.get('FLASK_ENV', 'default')].init_app(app)
db.init_app(app)

import json as _json
@app.template_filter('from_json')
def from_json_filter(s):
    try:
        return _json.loads(s) if s else []
    except Exception:
        return []

@app.teardown_appcontext
def cleanup_session(exception=None):
    if exception:
        db.session.rollback()

with app.app_context():
    db.create_all()
    # 自動遷移：補上新欄位
    from sqlalchemy import text, inspect
    with db.engine.connect() as conn:
        insp = inspect(db.engine)
        # Products 表新增欄位
        prod_cols = [c['name'] for c in insp.get_columns('products')]
        for col_name, col_def in [
            ('ko_lockout', 'INTEGER DEFAULT 1'),
            ('ko_start_pct', 'FLOAT'),
            ('ko_stepdown_pct', 'FLOAT'),
            ('observation_dates', 'TEXT'),
        ]:
            if col_name not in prod_cols:
                conn.execute(text(f'ALTER TABLE products ADD COLUMN {col_name} {col_def}'))
                conn.commit()
                app.logger.info(f'已自動新增 products.{col_name} 欄位')
        # Underlyings 表新增欄位
        ul_cols = [c['name'] for c in insp.get_columns('underlyings')]
        if 'ko_hit_date' not in ul_cols:
            conn.execute(text('ALTER TABLE underlyings ADD COLUMN ko_hit_date DATE'))
            conn.commit()
            app.logger.info('已自動新增 underlyings.ko_hit_date 欄位')
        # Clients 表新增 B 模組欄位（客戶分群+個人化）
        client_cols = [c['name'] for c in insp.get_columns('clients')]
        for col_name, col_def in [
            ('risk_profile', 'VARCHAR(20)'),
            ('tags', 'TEXT'),
            ('notes', 'TEXT'),
        ]:
            if col_name not in client_cols:
                conn.execute(text(f'ALTER TABLE clients ADD COLUMN {col_name} {col_def}'))
                conn.commit()
                app.logger.info(f'已自動新增 clients.{col_name} 欄位')
        # PostgreSQL: 修正 ID 序列不同步
        db_uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if 'postgresql' in db_uri:
            for tbl in ['products', 'underlyings', 'clients', 'positions', 'app_users']:
                try:
                    conn.execute(text(f"SELECT setval('{tbl}_id_seq', COALESCE((SELECT MAX(id) FROM {tbl}), 1))"))
                    conn.commit()
                except Exception:
                    pass
            app.logger.info('已修正 PostgreSQL ID 序列')

# 一次性資料修正：統一 2026 年 active 商品的 special_notes（發行日｜到期日｜ISIN）、
# 修正 2026SN2301 比價日 = 2026-06-15、補建 2026SN2301 配息排程（套用完成後可移除此區塊）
with app.app_context():
    try:
        from datetime import date as _date
        from models import PaymentSchedule
        _UNIFY = {
            '2026SN1891': ('發行日 2026-04-30｜到期日 2026-11-03｜ISIN XS3342478644', None),
            '2026SN1495': ('發行日 2026-04-22｜到期日 2027-01-27｜ISIN XS3263839709', None),
            '2026SN1415': ('發行日 2026-04-20｜到期日 2026-09-24｜ISIN XS3326009233', None),
            '2026SN2301': ('發行日 2026-05-15｜到期日 2027-05-19｜ISIN XS3365601874', _date(2026, 6, 15)),
        }
        _changed = False
        for _code, (_notes, _sd) in _UNIFY.items():
            for _p in Product.query.filter_by(product_code=_code).all():
                if _p.special_notes != _notes:
                    _p.special_notes = _notes
                    _changed = True
                if _sd and _p.start_date != _sd:
                    _p.start_date = _sd
                    _changed = True
        # 補建配息排程（照 TS 表A）— code: [(period, obs_start, obs_end, payment_date), ...]
        _SCHEDULES = {
            '2026SN2301': [
                (1,  '2026-05-11', '2026-06-15', '2026-06-17'),
                (2,  '2026-06-16', '2026-07-15', '2026-07-17'),
                (3,  '2026-07-16', '2026-08-17', '2026-08-19'),
                (4,  '2026-08-18', '2026-09-15', '2026-09-17'),
                (5,  '2026-09-16', '2026-10-15', '2026-10-19'),
                (6,  '2026-10-16', '2026-11-16', '2026-11-18'),
                (7,  '2026-11-17', '2026-12-15', '2026-12-17'),
                (8,  '2026-12-16', '2027-01-15', '2027-01-20'),
                (9,  '2027-01-16', '2027-02-16', '2027-02-18'),
                (10, '2027-02-17', '2027-03-15', '2027-03-17'),
                (11, '2027-03-16', '2027-04-15', '2027-04-19'),
                (12, '2027-04-16', '2027-05-17', '2027-05-19'),
            ],
            '2026SN2155': [
                (1, '2026-05-06', '2026-06-12', '2026-06-16'),
                (2, '2026-06-13', '2026-07-13', '2026-07-15'),
                (3, '2026-07-14', '2026-08-12', '2026-08-14'),
                (4, '2026-08-13', '2026-09-14', '2026-09-16'),
                (5, '2026-09-15', '2026-10-12', '2026-10-14'),
                (6, '2026-10-13', '2026-11-12', '2026-11-16'),
                (7, '2026-11-13', '2026-12-14', '2026-12-16'),
                (8, '2026-12-15', '2027-01-12', '2027-01-14'),
            ],
            '2026SN2417': [
                (1,  '2026-05-13', '2026-06-22', '2026-06-24'),
                (2,  '2026-06-23', '2026-07-20', '2026-07-22'),
                (3,  '2026-07-21', '2026-08-19', '2026-08-21'),
                (4,  '2026-08-20', '2026-09-21', '2026-09-23'),
                (5,  '2026-09-22', '2026-10-19', '2026-10-21'),
                (6,  '2026-10-20', '2026-11-19', '2026-11-23'),
                (7,  '2026-11-20', '2026-12-21', '2026-12-23'),
                (8,  '2026-12-22', '2027-01-19', '2027-01-21'),
                (9,  '2027-01-20', '2027-02-19', '2027-02-23'),
                (10, '2027-02-20', '2027-03-19', '2027-03-23'),
                (11, '2027-03-20', '2027-04-19', '2027-04-21'),
                (12, '2027-04-20', '2027-05-19', '2027-05-21'),
            ],
            '2026SN1650': [
                (1, '2026-04-20', '2026-05-26', '2026-05-28'),
                (2, '2026-05-27', '2026-06-24', '2026-06-26'),
                (3, '2026-06-25', '2026-07-24', '2026-07-28'),
                (4, '2026-07-25', '2026-08-24', '2026-08-26'),
                (5, '2026-08-25', '2026-09-24', '2026-09-28'),
                (6, '2026-09-25', '2026-10-26', '2026-10-28'),
                (7, '2026-10-27', '2026-11-24', '2026-11-27'),
                (8, '2026-11-25', '2026-12-24', '2026-12-29'),
            ],
            '2026SN1938': [
                (1, '2026-05-07', '2026-06-08', '2026-06-11'),
                (2, '2026-06-09', '2026-07-07', '2026-07-10'),
                (3, '2026-07-08', '2026-08-07', '2026-08-12'),
                (4, '2026-08-10', '2026-09-08', '2026-09-11'),
                (5, '2026-09-09', '2026-10-07', '2026-10-13'),
                (6, '2026-10-08', '2026-11-09', '2026-11-13'),
            ],
        }
        for _code, _sched in _SCHEDULES.items():
            for _p in Product.query.filter_by(product_code=_code).all():
                if not _p.payment_schedule:
                    for _per, _os, _oe, _pd in _sched:
                        db.session.add(PaymentSchedule(
                            product_id=_p.id, period=_per,
                            obs_start_date=_date.fromisoformat(_os),
                            obs_end_date=_date.fromisoformat(_oe),
                            payment_date=_date.fromisoformat(_pd)))
                    _changed = True
        if _changed:
            db.session.commit()
            app.logger.info('已套用 2026 商品資料統一修正')
    except Exception as _e:
        db.session.rollback()
        app.logger.warning(f'2026 商品資料統一修正失敗: {_e}')

# ── 初始化 LINE Bot ──────────────────────────────────────────────────────────
from line_bot import init_line, is_configured as line_is_configured
init_line(app)

# ── 排程：每日自動發送市場日報 ───────────────────────────────────────────────
def _setup_scheduler():
    """設定排程（避免 gunicorn 多 worker 重複啟動）"""
    try:
        from apscheduler.schedulers.background import BackgroundScheduler

        def _scheduled_daily_report():
            from line_bot import broadcast_text, is_configured
            if not is_configured():
                return
            with app.app_context():
                from daily_report import generate_market_report, generate_settlement_alert
                # 訊息 A：市場日報（純市場，每天發）
                report = generate_market_report(app)
                broadcast_text(report)
                app.logger.info('每日市場日報已自動發送')
                # 訊息 B：商品異動提醒（事件驅動，沒事不發）
                alert = generate_settlement_alert(app)
                if alert:
                    broadcast_text(alert)
                    app.logger.info('商品異動提醒已自動發送')

        def _scheduled_price_update():
            with app.app_context():
                # 部署重啟後舊連線可能已斷，先清除再建新連線
                db.session.remove()
                from price_fetcher import fetch_quotes
                from dateutil.relativedelta import relativedelta
                users = AppUser.query.all()
                today = date.today()
                for user in users:
                    try:
                        active = Product.query.filter_by(status='active', user_id=user.id).all()
                        tickers = set()
                        for p in active:
                            for u in p.underlyings:
                                if u.ticker:
                                    tickers.add(u.ticker)
                        if not tickers:
                            continue
                        try:
                            prices, price_date = fetch_quotes(tickers)
                        except Exception as e:
                            app.logger.error(f'排程 fetch_quotes user {user.id} 例外: {e}')
                            prices, price_date = {}, None
                        if not prices:
                            app.logger.warning(f'排程更新 user {user.id}: 所有來源皆失敗，0 檔更新')
                            continue
                        if not price_date:
                            price_date = today - timedelta(days=1)
                        for p in active:
                            for u in p.underlyings:
                                if not u.ticker:
                                    continue
                                try:
                                    if u.ticker in prices:
                                        u.latest_price = prices[u.ticker]
                                        u.price_date = price_date
                                    # ── Stepdown FCN: 比價日自動判定 ──
                                    if p.ko_type == 'stepdown' and not u.ko_hit and p.observation_dates and p.ko_start_pct and p.ko_stepdown_pct:
                                        obs_list = _json.loads(p.observation_dates) if isinstance(p.observation_dates, str) else []
                                        for idx, obs_str in enumerate(obs_list):
                                            try:
                                                obs_d = datetime.strptime(obs_str, '%Y-%m-%d').date()
                                            except Exception:
                                                continue
                                            if obs_d != price_date:
                                                continue
                                            ko_pct_m = p.ko_start_pct - p.ko_stepdown_pct * (idx + 1)
                                            ko_price_m = u.initial_price * ko_pct_m if u.initial_price else None
                                            if ko_price_m and u.latest_price and u.latest_price >= ko_price_m:
                                                u.ko_hit = True
                                                u.ko_hit_date = obs_d
                                                app.logger.info(f'[排程] Stepdown KO 鎖定: {p.product_code} {u.ticker} 月{idx+1}')
                                            break
                                    # ── 一般 FCN: start_date = 比價日 = KO 觀察起始日 ──
                                    elif p.ko_type != 'stepdown' and u.ko_level and u.latest_price and p.start_date:
                                        if today >= p.start_date:
                                            if u.latest_price >= u.ko_level:
                                                u.ko_hit = True
                                except Exception:
                                    db.session.rollback()
                        db.session.commit()
                    except Exception as e:
                        db.session.rollback()
                        app.logger.error(f'排程更新 user {user.id} 失敗: {e}')
                app.logger.info('每日收盤價已自動更新')
                # 備份資料庫到 Google Drive
                try:
                    from gdrive_backup import backup_to_drive
                    backup_to_drive(app)
                except Exception as e:
                    app.logger.error(f'自動備份失敗: {e}')

        def _keep_alive():
            """每 10 分鐘 ping 自己，防止 Render 免費方案休眠"""
            try:
                import requests as _req
                url = os.environ.get('RENDER_EXTERNAL_URL', 'https://fcnmanager.onrender.com')
                _req.get(f'{url}/health', timeout=10)
            except Exception:
                pass

        scheduler = BackgroundScheduler()
        scheduler.add_job(_scheduled_price_update, 'cron', hour=6, minute=0,
                          timezone='Asia/Taipei', id='daily_price_update')
        scheduler.add_job(_scheduled_daily_report, 'cron', hour=7, minute=30,
                          timezone='Asia/Taipei', id='daily_report')
        # 啟動後 30 秒執行一次更新（Render 重啟後馬上補抓）
        scheduler.add_job(_scheduled_price_update, 'date',
                          run_date=datetime.now() + timedelta(seconds=30),
                          id='startup_price_update')
        # 每 10 分鐘 ping 自己，防止休眠
        scheduler.add_job(_keep_alive, 'interval', minutes=10, id='keep_alive')
        scheduler.start()
        return scheduler
    except Exception as e:
        app.logger.warning(f'排程啟動失敗: {e}')
        return None

scheduler = _setup_scheduler()

# ── 初始化預設帳號 ────────────────────────────────────────────────────────────
with app.app_context():
    if not AppUser.query.first():
        u = AppUser(username='admin', role='admin')
        u.set_password('fcn2026')
        db.session.add(u)
        db.session.commit()


# ── 登入驗證 ─────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


@app.route('/health')
def health():
    return 'ok', 200


@app.route('/debug/prices')
def debug_prices():
    """診斷用：測試各報價來源在 Render 上是否能正常運作"""
    import json
    from price_fetcher import _QUOTE_SOURCES
    today = date.today()
    results = []
    test_tickers = ['AAPL']
    for fn in _QUOTE_SOURCES[:5]:
        try:
            prices, pdate = fn(test_tickers, today)
            results.append({'source': fn.__name__, 'prices': prices,
                           'date': str(pdate) if pdate else None, 'error': None})
        except Exception as e:
            results.append({'source': fn.__name__, 'prices': {},
                           'date': None, 'error': str(e)})
    return json.dumps(results, indent=2), 200, {'Content-Type': 'application/json'}


@app.after_request
def optimize_response(response):
    # 靜態檔案快取 1 天
    if request.path.startswith('/static/'):
        response.cache_control.max_age = 86400
        response.cache_control.public = True
    return response


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = AppUser.query.filter_by(username=request.form['username']).first()
        if user and user.check_password(request.form['password']):
            session['logged_in'] = True
            session['user_id'] = user.id
            session['username'] = user.username
            session['is_admin'] = user.is_admin
            log_activity('登入')
            return redirect(url_for('dashboard'))
        flash('帳號或密碼錯誤', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))


@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        user = AppUser.query.get(session['user_id'])
        old_pw = request.form['old_password']
        new_pw = request.form['new_password']
        confirm_pw = request.form['confirm_password']
        if not user.check_password(old_pw):
            flash('舊密碼錯誤', 'danger')
        elif new_pw != confirm_pw:
            flash('新密碼與確認密碼不一致', 'danger')
        elif len(new_pw) < 4:
            flash('密碼至少 4 個字元', 'danger')
        else:
            user.set_password(new_pw)
            db.session.commit()
            flash('密碼已更新', 'success')
            return redirect(url_for('dashboard'))
    return render_template('change_password.html')


TICKER_NAME = {
    'NVDA': '輝達', 'AVGO': '博通', 'TSM': '台積電', 'AMD': '超微',
    'MU': '美光', 'ARM': '安謀', 'QCOM': '高通', 'INTC': '英特爾',
    'AAPL': '蘋果', 'MSFT': '微軟', 'GOOG': '谷歌', 'GOOGL': '谷歌', 'AMZN': '亞馬遜',
    'META': '臉書', 'TSLA': '特斯拉', 'ORCL': '甲骨文', 'CRM': '賽富時',
    'NFLX': '網飛', 'UBER': '優步', 'UNH': '聯合健康', 'JPM': '摩根大通',
    'GS': '高盛', 'BA': '波音', 'AAL': '美國航空', 'AA': '美國鋁業',
    'NKE': '耐吉', 'COIN': '幣安所', 'SMCI': '超微電腦', 'ASML': '艾司摩爾',
    'VST': '維斯達公司', 'F': '福特', 'DIS': '迪士尼', 'PYPL': '貝寶',
    'CCL': '嘉年華', 'X': '美國鋼鐵', 'SOFI': '索飛', 'PLTR': '帕蘭提爾',
    'MSTR': '微策略', 'UAL': '聯合航空', 'CRWV': 'CrowdStrike',
    'NEM': '紐蒙特礦業', 'BMNR': 'Beamr Imaging',
}

@app.context_processor
def inject_globals():
    return {'TICKER_NAME': TICKER_NAME, 'is_admin': session.get('is_admin', False)}


def current_uid():
    return session.get('user_id')


def log_activity(action):
    """記錄使用者操作"""
    try:
        uid = session.get('user_id')
        uname = session.get('username', '')
        if not uname and uid:
            u = AppUser.query.get(uid)
            uname = u.username if u else ''
        entry = ActivityLog(user_id=uid, username=uname, action=action,
                            ip=request.remote_addr)
        db.session.add(entry)
        db.session.commit()
    except Exception:
        pass


# ── 使用者管理（僅管理員）─────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('is_admin'):
            flash('需要管理員權限', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


@app.route('/users')
@login_required
@admin_required
def users():
    all_users = AppUser.query.order_by(AppUser.created_at).all()
    return render_template('users.html', users=all_users)


@app.route('/users/add', methods=['POST'])
@login_required
@admin_required
def add_user():
    username = request.form['username'].strip()
    password = request.form['password'].strip()
    if not username or not password:
        flash('帳號和密碼不可空白', 'danger')
        return redirect(url_for('users'))
    if AppUser.query.filter_by(username=username).first():
        flash('此帳號已存在', 'warning')
        return redirect(url_for('users'))
    u = AppUser(username=username, role='user')
    u.set_password(password)
    db.session.add(u)
    db.session.commit()
    flash(f'已新增使用者：{username}', 'success')
    return redirect(url_for('users'))


@app.route('/users/<int:uid>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(uid):
    u = AppUser.query.get_or_404(uid)
    if u.is_admin:
        flash('無法刪除管理員帳號', 'danger')
    else:
        db.session.delete(u)
        db.session.commit()
        flash(f'已刪除使用者：{u.username}', 'success')
    return redirect(url_for('users'))


@app.route('/users/<int:uid>/reset', methods=['POST'])
@login_required
@admin_required
def reset_user_password(uid):
    u = AppUser.query.get_or_404(uid)
    new_pw = request.form['new_password'].strip()
    if not new_pw:
        flash('密碼不可空白', 'danger')
    else:
        u.set_password(new_pw)
        db.session.commit()
        flash(f'已重設 {u.username} 的密碼', 'success')
    return redirect(url_for('users'))


# ── 匯出 Excel ───────────────────────────────────────────────────────────────
@app.route('/export_excel')
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    thin = Border(
        left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0'),
    )
    # 表頭：深橄欖綠 + 白字
    header_fill = PatternFill(start_color='4F6228', end_color='4F6228', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF', size=10)
    # 標的列 ticker 黃底
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    # KO 主資料行淺灰底
    info_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    # 標籤欄底色（金橘色）
    label_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    label_font = Font(bold=True, size=9, color='333333')
    # 交替底色
    alt_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')  # 淺藍
    # 一般字型
    normal_font = Font(size=9)
    bold_font = Font(bold=True, size=9)
    pct_font = Font(size=9, color='C00000')
    pct_fmt = '0.00%'
    date_fmt = 'YYYY/M/D'

    col_widths = {'A':7, 'B':13.4, 'C':9.7, 'D':5.9, 'E':10, 'F':10.4,
                  'G':8.8, 'H':10.4, 'I':12.1, 'J':11.1, 'K':10.4,
                  'L':9.3, 'M':9.6, 'N':9.3, 'O':9.6, 'P':11.4,
                  'Q':13, 'R':13, 'S':10.6, 'T':12.3}

    def write_sheet(ws, products, is_ko_sheet=False):
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        # 表頭
        headers = ['END' if is_ko_sheet else '', '商品代碼', '類別', 'Tenor',
                   '交易營業員', '', '設定%', '(V/X)', 'Underlying 1',
                   '(V/X)', 'Underlying 2', '(V/X)', 'Underlying 3',
                   '(V/X)', 'Underlying 4', '訂價日', '比價日',
                   '期末訂價日', '年利率', '']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='center')

        row = 2
        prod_idx = 0
        for p in products:
            uls = sorted(p.underlyings, key=lambda u: u.position_order or 0)
            clients = []
            for pos in p.positions:
                amt = ''
                if pos.investment_amount:
                    amt = str(int(pos.investment_amount / 10000)) + '萬'
                clients.append(pos.client.name_masked + amt)

            # 交替底色
            bg = alt_fill if prod_idx % 2 == 1 else None

            # 整個區塊先上底色
            if bg:
                for r in range(row, row + 5):
                    for col in range(1, 21):
                        ws.cell(row=r, column=col).fill = bg

            # ── Stepdown: 解析比價日 ──
            obs_dates = []
            if p.ko_type == 'stepdown' and p.observation_dates:
                import json as _j
                try:
                    obs_dates = _j.loads(p.observation_dates)
                except Exception:
                    pass

            # ── Row 1: 標的 ──
            # E 欄: stepdown 或幣別
            if p.ko_type == 'stepdown' and p.ko_stepdown_pct:
                sd_label = f'stepdown {p.ko_stepdown_pct:.0%}'
                if p.currency and p.currency != 'USD':
                    ws.cell(row=row, column=5, value=p.currency).font = normal_font
                    ws.cell(row=row+1, column=5, value=sd_label).font = normal_font
                else:
                    ws.cell(row=row, column=5, value=sd_label).font = normal_font
            elif p.currency and p.currency != 'USD':
                ws.cell(row=row, column=5, value=p.currency).font = normal_font

            c = ws.cell(row=row, column=6, value='標的')
            c.border = thin; c.font = label_font; c.fill = label_fill
            for i, u in enumerate(uls[:4]):
                c = ws.cell(row=row, column=9 + i*2, value=u.ticker)
                c.border = thin; c.fill = yellow_fill; c.font = bold_font
            if len(clients) > 0:
                ws.cell(row=row, column=20, value=clients[0]).border = thin

            # ── Row 2: 期初價格 ──
            # E 欄: 比價日（前半）
            if obs_dates:
                half = len(obs_dates) // 2
                dates_1 = '、'.join(d[5:].replace('-', '/') for d in obs_dates[:half])
                dates_2 = '、'.join(d[5:].replace('-', '/') for d in obs_dates[half:])
                ws.cell(row=row+1, column=5, value=dates_1).font = normal_font

            c = ws.cell(row=row+1, column=6, value='期初價格')
            c.border = thin; c.font = label_font; c.fill = label_fill
            for i, u in enumerate(uls[:4]):
                c = ws.cell(row=row+1, column=9 + i*2, value=u.initial_price)
                c.border = thin; c.font = normal_font
            if len(clients) > 1:
                ws.cell(row=row+1, column=20, value=clients[1]).border = thin

            # ── Row 3: KO（主資料行）──
            # E 欄: 比價日（後半）
            if obs_dates:
                ws.cell(row=row+2, column=5, value=dates_2).font = normal_font

            c = ws.cell(row=row+2, column=2, value=p.product_code)
            c.border = thin; c.font = bold_font
            c = ws.cell(row=row+2, column=3, value=p.product_type or 'FCN')
            c.border = thin; c.font = normal_font
            c = ws.cell(row=row+2, column=4, value=p.tenor_months)
            c.border = thin; c.font = normal_font
            c = ws.cell(row=row+2, column=6, value='KO')
            c.border = thin; c.font = label_font; c.fill = label_fill
            # Stepdown: KO 欄位用公式格式顯示
            if p.ko_type == 'stepdown' and p.ko_start_pct and p.ko_stepdown_pct:
                ko_display = p.ko_pct
                c = ws.cell(row=row+2, column=7, value=ko_display)
            else:
                c = ws.cell(row=row+2, column=7, value=p.ko_pct)
            c.border = thin; c.number_format = pct_fmt; c.font = pct_font; c.fill = label_fill
            for i, u in enumerate(uls[:4]):
                c = ws.cell(row=row+2, column=8 + i*2, value='V' if u.ko_hit else '')
                c.border = thin; c.font = normal_font
                c = ws.cell(row=row+2, column=9 + i*2, value=u.ko_level)
                c.border = thin; c.font = normal_font
            if p.trade_date:
                c = ws.cell(row=row+2, column=16, value=p.trade_date)
                c.border = thin; c.number_format = date_fmt; c.font = normal_font
            if p.start_date:
                c = ws.cell(row=row+2, column=17, value=p.start_date)
                c.border = thin; c.number_format = date_fmt; c.font = normal_font
            if p.maturity_date:
                c = ws.cell(row=row+2, column=18, value=p.maturity_date)
                c.border = thin; c.number_format = date_fmt; c.font = normal_font
            if p.coupon_rate:
                c = ws.cell(row=row+2, column=19, value=p.coupon_rate)
                c.border = thin; c.number_format = pct_fmt; c.font = pct_font
            if len(clients) > 2:
                ws.cell(row=row+2, column=20, value=clients[2]).border = thin

            # ── Row 4: Strike ──
            c = ws.cell(row=row+3, column=6, value='Strike')
            c.border = thin; c.font = label_font; c.fill = label_fill
            c = ws.cell(row=row+3, column=7, value=p.strike_pct)
            c.border = thin; c.number_format = pct_fmt; c.font = pct_font; c.fill = label_fill
            for i, u in enumerate(uls[:4]):
                c = ws.cell(row=row+3, column=9 + i*2, value=u.strike_level)
                c.border = thin; c.font = normal_font
            if len(clients) > 3:
                ws.cell(row=row+3, column=20, value=clients[3]).border = thin

            # ── Row 5: EKI ──
            has_eki = p.eki_pct and p.eki_pct > 0
            c = ws.cell(row=row+4, column=6, value='EKI' if has_eki else '無EKI')
            c.border = thin; c.font = label_font; c.fill = label_fill
            c = ws.cell(row=row+4, column=7, value=p.eki_pct if has_eki else 0)
            c.border = thin; c.number_format = pct_fmt; c.font = pct_font; c.fill = label_fill
            for i, u in enumerate(uls[:4]):
                c = ws.cell(row=row+4, column=9 + i*2, value=u.eki_level if has_eki else 0)
                c.border = thin; c.font = normal_font

            row += 5
            prod_idx += 1

    # Sheet 1: 持倉
    ws1 = wb.active
    ws1.title = '持倉'
    active = Product.query.filter_by(status='active', user_id=current_uid()).order_by(Product.trade_date, Product.created_at).all()
    write_sheet(ws1, active)

    # Sheet 2: KO
    ws2 = wb.create_sheet('KO')
    ko = Product.query.filter_by(status='ko_exited', user_id=current_uid()).order_by(Product.created_at).all()
    write_sheet(ws2, ko, is_ko_sheet=True)

    # Sheet 3: 到期
    ws3 = wb.create_sheet('到期')
    matured = Product.query.filter_by(status='matured', user_id=current_uid()).order_by(Product.created_at).all()
    write_sheet(ws3, matured)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name=f'SN記憶表格_{date.today().strftime("%Y%m%d")}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 背景靜默更新股價（使用者無感） ──────────────────────────────────────────
import threading

_bg_lock = threading.Lock()
_bg_running = False

def _maybe_bg_update(uid, active):
    """價格過期時，背景執行緒靜默更新，不阻塞頁面載入"""
    global _bg_running
    if _bg_running:
        return

    today = date.today()
    # 判斷是否有任何標的的價格日期 < 上一個交易日
    needs_update = False
    for p in active:
        for u in p.underlyings:
            if u.ticker and (not u.price_date or u.price_date < today - timedelta(days=1)):
                needs_update = True
                break
        if needs_update:
            break

    if not needs_update:
        return

    def _do_update():
        global _bg_running
        with _bg_lock:
            _bg_running = True
        try:
            with app.app_context():
                db.session.remove()
                from price_fetcher import fetch_quotes
                from dateutil.relativedelta import relativedelta
                all_products = Product.query.filter_by(status='active', user_id=uid).all()
                tickers = set()
                for p in all_products:
                    for u in p.underlyings:
                        if u.ticker:
                            tickers.add(u.ticker)
                if not tickers:
                    return

                try:
                    prices, price_date = fetch_quotes(tickers)
                except Exception:
                    prices, price_date = {}, None

                if not prices:
                    return
                if not price_date:
                    price_date = today - timedelta(days=1)

                for p in all_products:
                    for u in p.underlyings:
                        if not u.ticker:
                            continue
                        try:
                            if u.ticker in prices:
                                u.latest_price = prices[u.ticker]
                                u.price_date = price_date
                            # 一般 FCN: start_date = 比價日 = KO 觀察起始日
                            if p.ko_type != 'stepdown' and u.ko_level and u.latest_price and p.start_date:
                                if today >= p.start_date:
                                    if u.latest_price >= u.ko_level:
                                        u.ko_hit = True
                        except Exception:
                            db.session.rollback()

                try:
                    db.session.commit()
                    app.logger.info(f'背景靜默更新完成：{len(prices)} 檔')
                except Exception:
                    db.session.rollback()
        finally:
            with _bg_lock:
                _bg_running = False

    threading.Thread(target=_do_update, daemon=True).start()


# ── 首頁：持倉總覽 ────────────────────────────────────────────────────────────
@app.route('/')
@login_required
def dashboard():
    try:
        active = Product.query.filter_by(status='active', user_id=current_uid()) \
            .options(joinedload(Product.underlyings), joinedload(Product.positions).joinedload(Position.client)) \
            .order_by(Product.trade_date, Product.created_at).all()
    except Exception as e:
        app.logger.error(f'Dashboard query error: {e}')
        active = Product.query.filter_by(status='active', user_id=current_uid()).all()

    # 一般 FCN: 根據最新收盤價自動更新 KO 狀態
    # start_date = 比價日 = KO 觀察起始日（含當日）
    today = date.today()
    ko_changed = False
    for p in active:
        if p.ko_type != 'stepdown' and p.start_date and today >= p.start_date:
            for u in p.underlyings:
                if u.ko_level and u.latest_price and not u.ko_hit:
                    if u.latest_price >= u.ko_level:
                        u.ko_hit = True
                        ko_changed = True
    if ko_changed:
        db.session.commit()

    # 全部達 KO 只提醒，不自動移出場（由使用者自行決定）

    # 檢查價格是否過期，過期就背景靜默更新（不阻塞頁面載入）
    _maybe_bg_update(current_uid(), active)

    # 為 Stepdown FCN 預算月度排程
    stepdown_schedules = {}
    for p in active:
        if p.ko_type == 'stepdown' and p.observation_dates and p.ko_start_pct and p.ko_stepdown_pct:
            obs = _json.loads(p.observation_dates) if isinstance(p.observation_dates, str) else []
            schedule = []
            for i, obs_date_str in enumerate(obs):
                ko_pct_month = p.ko_start_pct - p.ko_stepdown_pct * (i + 1)
                try:
                    obs_date = datetime.strptime(obs_date_str, '%Y-%m-%d').date()
                except Exception:
                    obs_date = None
                # 每檔標的在這個月的狀態
                ul_status = {}
                for u in p.underlyings:
                    if u.ko_hit and u.ko_hit_date and obs_date and u.ko_hit_date <= obs_date:
                        ul_status[u.id] = 'locked'  # 已鎖定（之前月份通過）
                    elif u.ko_hit and u.ko_hit_date and obs_date and u.ko_hit_date == obs_date:
                        ul_status[u.id] = 'hit'     # 本月通過
                    elif obs_date and obs_date < date.today():
                        ul_status[u.id] = 'miss'    # 已過未通過
                    elif obs_date and obs_date == date.today():
                        ul_status[u.id] = 'today'   # 今天比價
                    else:
                        ul_status[u.id] = 'pending'  # 未到
                schedule.append({
                    'month': i + 1,
                    'date': obs_date,
                    'date_str': obs_date_str[5:].replace('-', '/') if obs_date_str else '-',
                    'ko_pct': ko_pct_month,
                    'ul_status': ul_status,
                })
            stepdown_schedules[p.id] = schedule

    return render_template('dashboard.html', active=active, today=date.today(),
                           stepdown_schedules=stepdown_schedules)


# ── 已出場(KO)頁面 ───────────────────────────────────────────────────────────
@app.route('/ko')
@login_required
def ko_history():
    products = Product.query.filter_by(status='ko_exited', user_id=current_uid()).order_by(Product.maturity_date.desc()).all()
    return render_template('ko_history.html', products=products)


# ── 更新收盤價 ────────────────────────────────────────────────────────────────
@app.route('/fetch_prices')
@login_required
def fetch_prices():
    log_activity('更新收盤價')
    uid = current_uid()

    def _bg_fetch(uid):
        with app.app_context():
            db.session.remove()
            from price_fetcher import fetch_quotes
            from dateutil.relativedelta import relativedelta
            active = Product.query.filter_by(status='active', user_id=uid).all()
            tickers = set()
            for p in active:
                for u in p.underlyings:
                    if u.ticker:
                        tickers.add(u.ticker)
            if not tickers:
                return
            today = date.today()
            try:
                prices, price_date = fetch_quotes(tickers)
            except Exception as e:
                app.logger.error(f'手動 fetch_quotes user={uid} 例外: {e}')
                prices, price_date = {}, None
            if not prices:
                app.logger.warning(f'手動更新 user={uid}: 所有來源皆失敗，0 檔更新')
                return
            if not price_date:
                price_date = today - timedelta(days=1)
            for p in active:
                for u in p.underlyings:
                    if not u.ticker:
                        continue
                    try:
                        if u.ticker in prices:
                            u.latest_price = prices[u.ticker]
                            u.price_date = price_date
                            # 同時寫入 PriceHistory
                            existing = PriceHistory.query.filter_by(underlying_id=u.id, price_date=price_date).first()
                            if not existing:
                                db.session.add(PriceHistory(underlying_id=u.id, price_date=price_date, closing_price=prices[u.ticker]))
                        # ── Stepdown FCN: 比價日自動判定 ──
                        if p.ko_type == 'stepdown' and not u.ko_hit and p.observation_dates and p.ko_start_pct and p.ko_stepdown_pct:
                            obs_list = _json.loads(p.observation_dates) if isinstance(p.observation_dates, str) else []
                            for idx, obs_str in enumerate(obs_list):
                                try:
                                    obs_date = datetime.strptime(obs_str, '%Y-%m-%d').date()
                                except Exception:
                                    continue
                                if obs_date != price_date:
                                    continue
                                # 今天是比價日，計算該月 KO 價格
                                ko_pct_month = p.ko_start_pct - p.ko_stepdown_pct * (idx + 1)
                                ko_price_month = u.initial_price * ko_pct_month if u.initial_price else None
                                if ko_price_month and u.latest_price and u.latest_price >= ko_price_month:
                                    u.ko_hit = True
                                    u.ko_hit_date = obs_date
                                    app.logger.info(f'Stepdown KO 鎖定: {p.product_code} {u.ticker} 月{idx+1} ({obs_str}) 收盤{u.latest_price:.2f} >= KO{ko_price_month:.2f}')
                                break  # 一天只會匹配一個比價日

                        # ── 一般 FCN: start_date = 比價日 = KO 觀察起始日 ──
                        elif p.ko_type != 'stepdown' and u.ko_level and u.latest_price and p.start_date:
                            if today >= p.start_date:
                                if u.latest_price >= u.ko_level:
                                    u.ko_hit = True
                    except Exception:
                        db.session.rollback()
            try:
                db.session.commit()
                app.logger.info(f'手動更新完成：user={uid}')
            except Exception as e:
                db.session.rollback()
                app.logger.error(f'手動更新 DB commit 失敗 user={uid}: {e}')

    threading.Thread(target=_bg_fetch, args=(uid,), daemon=True).start()
    flash('正在背景更新收盤價，約 10 秒後重新整理即可看到最新價格', 'info')
    return redirect(url_for('dashboard'))


# ── 手動備份到 Google Drive ───────────────────────────────────────────────────
@app.route('/backup_drive')
@login_required
def backup_drive():
    try:
        from gdrive_backup import backup_to_drive
        result = backup_to_drive(app)
        if result:
            flash('資料庫已成功備份到 Google Drive', 'success')
        else:
            flash('備份失敗，請檢查 Google Drive 設定', 'danger')
    except Exception as e:
        flash(f'備份失敗: {e}', 'danger')
    return redirect(url_for('dashboard'))


# ── F 模組：FCN 反對處理庫 ────────────────────────────────────────────────────
@app.route('/objections')
@login_required
@admin_required
def objections():
    """FCN 銷售反對處理話術庫（admin 限定）"""
    import json as _jsonO
    from pathlib import Path
    data_path = Path(__file__).parent / 'data' / 'objections.json'
    try:
        data = _jsonO.loads(data_path.read_text(encoding='utf-8'))
    except Exception as e:
        flash(f'反對處理庫資料讀取失敗：{e}', 'danger')
        return redirect(url_for('dashboard'))

    return render_template(
        'objections.html',
        categories=data.get('categories', []),
        items=data.get('items', []),
        version=data.get('version', ''),
    )


# ── 客戶管理 ──────────────────────────────────────────────────────────────────
@app.route('/clients')
@login_required
def clients():
    all_clients = Client.query.filter_by(user_id=current_uid()).order_by(Client.name).all()
    return render_template('clients/index.html', clients=all_clients)


@app.route('/clients/add', methods=['GET', 'POST'])
@login_required
def add_client():
    if request.method == 'POST':
        name = request.form['name'].strip()
        if not name:
            flash('請輸入客戶姓名', 'danger')
            return redirect(url_for('add_client'))
        existing = Client.query.filter_by(name=name, user_id=current_uid()).first()
        if existing:
            flash('此客戶已存在', 'warning')
            return redirect(url_for('clients'))
        c = Client(name=name, name_masked=Client.mask_name(name), user_id=current_uid())
        db.session.add(c)
        db.session.commit()
        flash(f'已新增客戶：{name}', 'success')
        return redirect(url_for('clients'))
    return render_template('clients/add.html')


@app.route('/clients/<int:cid>/delete', methods=['POST'])
@login_required
def delete_client(cid):
    c = Client.query.get_or_404(cid)
    if c.positions:
        flash(f'{c.name_masked} 尚有部位紀錄，無法刪除', 'danger')
    else:
        db.session.delete(c)
        db.session.commit()
        flash(f'已刪除客戶 {c.name_masked}', 'success')
    return redirect(url_for('clients'))


# B 模組：預定義 tag 與風險屬性選項
CLIENT_TAGS = [
    '半導體', 'AI 科技', '美股成長', '高股息', '收息固收',
    '債券', '醫療生技', '能源原物料', '金融', 'ESG',
]
CLIENT_RISK_PROFILES = ['保守', '穩健', '積極']


@app.route('/clients/<int:cid>/edit', methods=['GET', 'POST'])
@login_required
def edit_client(cid):
    """編輯客戶分群資料：tags / risk_profile / notes（給 B 模組用）"""
    import json as _json3
    c = Client.query.get_or_404(cid)
    if c.user_id != current_uid():
        abort(403)

    if request.method == 'POST':
        # tags 從 multi-checkbox 來
        selected_tags = request.form.getlist('tags')
        valid_tags = [t for t in selected_tags if t in CLIENT_TAGS]
        c.tags = _json3.dumps(valid_tags, ensure_ascii=False) if valid_tags else None

        rp = request.form.get('risk_profile', '').strip()
        c.risk_profile = rp if rp in CLIENT_RISK_PROFILES else None

        notes = request.form.get('notes', '').strip()
        c.notes = notes or None

        db.session.commit()
        log_activity(f'編輯客戶分群：{c.name_masked} tags={valid_tags} risk={rp}')
        flash(f'已更新 {c.name_masked} 的分群資料', 'success')
        return redirect(url_for('clients'))

    return render_template(
        'clients/edit.html',
        client=c,
        all_tags=CLIENT_TAGS,
        all_risk_profiles=CLIENT_RISK_PROFILES,
        selected_tags=c.tags_list,
    )


# ── 商品管理 ──────────────────────────────────────────────────────────────────
@app.route('/products')
@login_required
def products():
    active = Product.query.filter_by(status='active', user_id=current_uid()).order_by(Product.trade_date, Product.created_at).all()
    ko_done = Product.query.filter_by(status='ko_exited', user_id=current_uid()).order_by(Product.created_at.desc()).all()
    matured = Product.query.filter_by(status='matured', user_id=current_uid()).order_by(Product.created_at.desc()).all()
    # 計算持倉總金額（分幣別）
    from collections import defaultdict
    totals_by_currency = defaultdict(float)
    for p in active:
        for pos in p.positions:
            if pos.investment_amount:
                totals_by_currency[p.currency or 'USD'] += pos.investment_amount
    return render_template('products/index.html', active=active, ko_done=ko_done,
                           matured=matured, totals_by_currency=dict(totals_by_currency), today=date.today())


@app.route('/products/add', methods=['GET', 'POST'])
@login_required
def add_product():
    clients = Client.query.filter_by(user_id=current_uid()).order_by(Client.name).all()
    if request.method == 'POST':
      try:
        f = request.form
        ko_type = f.get('ko_type', 'fixed')

        # Stepdown FCN: 解析比價日 + 計算 ko_pct（第一個月的 KO）
        ko_start_pct = None
        ko_stepdown_pct = None
        obs_dates_json = None
        ko_pct_val = None

        if ko_type == 'stepdown':
            ko_start_pct = float(f['ko_start_pct']) / 100 if f.get('ko_start_pct') else None
            ko_stepdown_pct = float(f['ko_stepdown_pct']) / 100 if f.get('ko_stepdown_pct') else None
            # 收集比價日
            obs_dates = []
            tenor = int(f['tenor_months']) if f.get('tenor_months') else 0
            for i in range(tenor):
                d = f.get(f'obs_date_{i}')
                if d:
                    obs_dates.append(d)
            if obs_dates:
                obs_dates_json = _json.dumps(obs_dates)
            # ko_pct 存第一個月的 KO（用於標的 KO 價格計算基準）
            if ko_start_pct and ko_stepdown_pct:
                ko_pct_val = ko_start_pct - ko_stepdown_pct
        else:
            ko_pct_val = float(f['ko_pct']) / 100 if f.get('ko_pct') else None

        # 建立商品
        p = Product(
            user_id       = current_uid(),
            product_code  = f['product_code'].strip(),
            issuer        = f.get('issuer', '').strip(),
            tenor_months  = int(f['tenor_months']) if f.get('tenor_months') else None,
            currency      = f.get('currency', 'USD'),
            coupon_rate   = float(f['coupon_rate']) / 100 if f.get('coupon_rate') else None,
            trade_date    = _parse_date(f.get('trade_date')),
            start_date    = _parse_date(f.get('start_date')),
            maturity_date = _parse_date(f.get('maturity_date')),
            ko_pct        = ko_pct_val,
            strike_pct    = float(f['strike_pct']) / 100 if f.get('strike_pct') else None,
            eki_pct       = float(f['eki_pct']) / 100 if f.get('eki_pct') else None,
            ko_type       = ko_type,
            ko_start_pct  = ko_start_pct,
            ko_stepdown_pct = ko_stepdown_pct,
            observation_dates = obs_dates_json,
            special_notes = f.get('special_notes', '').strip(),
            status        = 'active',
        )
        db.session.add(p)
        db.session.flush()

        # 建立標的
        for i in range(1, 5):
            ticker = f.get(f'ticker_{i}', '').strip()
            if not ticker:
                continue
            init_p = float(f[f'init_price_{i}']) if f.get(f'init_price_{i}') else None
            u = Underlying(
                product_id    = p.id,
                ticker        = ticker.upper(),
                initial_price = init_p,
                ko_level      = round(init_p * p.ko_pct, 4) if init_p and p.ko_pct else None,
                strike_level  = round(init_p * p.strike_pct, 4) if init_p and p.strike_pct else None,
                eki_level     = round(init_p * p.eki_pct, 4) if init_p and p.eki_pct else None,
                ko_hit        = False,
                position_order= i,
            )
            db.session.add(u)

        # 建立部位
        client_id = f.get('client_id')
        amount    = f.get('investment_amount')
        if client_id and amount:
            pos = Position(
                client_id         = int(client_id),
                product_id        = p.id,
                investment_amount = float(amount),
                notes             = f.get('pos_notes', '').strip(),
            )
            db.session.add(pos)

        db.session.commit()
        flash(f'已新增商品：{p.product_code}', 'success')
        return redirect(url_for('dashboard'))
      except Exception as e:
        db.session.rollback()
        app.logger.error(f'新增商品失敗: {e}')
        flash(f'新增失敗: {e}', 'danger')
        return redirect(url_for('add_product'))

    return render_template('products/add.html', clients=clients)


@app.route('/products/<int:pid>')
@login_required
def product_detail(pid):
    p = Product.query.get_or_404(pid)
    return render_template('products/detail.html', product=p, today=date.today())


@app.route('/products/<int:pid>/ko_exit', methods=['POST'])
@login_required
def ko_exit(pid):
    p = Product.query.get_or_404(pid)
    p.status = 'ko_exited'
    db.session.commit()
    flash(f'{p.product_code} 已標記為提前出場', 'success')
    return redirect(url_for('dashboard'))


@app.route('/products/<int:pid>/delete', methods=['POST'])
@login_required
def delete_product(pid):
    p = Product.query.get_or_404(pid)
    # 刪除相關標的、持倉
    for u in p.underlyings:
        db.session.delete(u)
    for pos in p.positions:
        db.session.delete(pos)
    db.session.delete(p)
    db.session.commit()
    flash(f'{p.product_code} 已刪除', 'success')
    return redirect(url_for('products'))


@app.route('/products/<int:pid>/reactivate', methods=['POST'])
@login_required
def reactivate(pid):
    p = Product.query.get_or_404(pid)
    p.status = 'active'
    db.session.commit()
    flash(f'{p.product_code} 已移回持倉中', 'success')
    return redirect(url_for('ko_history'))


@app.route('/products/<int:pid>/toggle_ko/<int:uid>', methods=['POST'])
@login_required
def toggle_ko(pid, uid):
    u = Underlying.query.get_or_404(uid)
    u.ko_hit = not u.ko_hit
    db.session.commit()
    return jsonify({'ko_hit': u.ko_hit})


# ── 編輯商品 ─────────────────────────────────────────────────────────────────
@app.route('/products/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
def edit_product(pid):
    p = Product.query.get_or_404(pid)
    if request.method == 'POST':
        f = request.form
        p.product_code  = f['product_code'].strip()
        p.issuer        = f.get('issuer', '').strip()
        p.tenor_months  = int(f['tenor_months']) if f.get('tenor_months') else None
        p.currency      = f.get('currency', 'USD')
        p.coupon_rate   = float(f['coupon_rate']) / 100 if f.get('coupon_rate') else None
        p.trade_date    = _parse_date(f.get('trade_date'))
        p.start_date    = _parse_date(f.get('start_date'))
        p.maturity_date = _parse_date(f.get('maturity_date'))
        p.ko_type       = f.get('ko_type', 'fixed')
        p.strike_pct    = float(f['strike_pct']) / 100 if f.get('strike_pct') else None
        p.eki_pct       = float(f['eki_pct']) / 100 if f.get('eki_pct') else None
        p.special_notes = f.get('special_notes', '').strip()

        if p.ko_type == 'stepdown':
            p.ko_start_pct = float(f['ko_start_pct']) / 100 if f.get('ko_start_pct') else None
            p.ko_stepdown_pct = float(f['ko_stepdown_pct']) / 100 if f.get('ko_stepdown_pct') else None
            # 收集比價日
            obs_dates = []
            tenor = int(f['tenor_months']) if f.get('tenor_months') else 0
            for i in range(tenor):
                d = f.get(f'obs_date_{i}')
                if d:
                    obs_dates.append(d)
            p.observation_dates = _json.dumps(obs_dates) if obs_dates else None
            # ko_pct 存第一個月 KO
            if p.ko_start_pct and p.ko_stepdown_pct:
                p.ko_pct = p.ko_start_pct - p.ko_stepdown_pct
        else:
            p.ko_pct = float(f['ko_pct']) / 100 if f.get('ko_pct') else None
            p.ko_start_pct = None
            p.ko_stepdown_pct = None
            p.observation_dates = None

        # 更新標的
        existing = {u.id: u for u in p.underlyings}
        for i in range(1, 5):
            uid = f.get(f'underlying_id_{i}')
            ticker = f.get(f'ticker_{i}', '').strip()
            init_p = float(f[f'init_price_{i}']) if f.get(f'init_price_{i}') else None
            if uid and int(uid) in existing:
                u = existing[int(uid)]
                if not ticker:
                    db.session.delete(u)
                    continue
                u.ticker = ticker.upper()
                u.initial_price = init_p
                u.ko_level = round(init_p * p.ko_pct, 4) if init_p and p.ko_pct else None
                u.strike_level = round(init_p * p.strike_pct, 4) if init_p and p.strike_pct else None
                u.eki_level = round(init_p * p.eki_pct, 4) if init_p and p.eki_pct else None
            elif ticker:
                u = Underlying(
                    product_id=p.id, ticker=ticker.upper(), initial_price=init_p,
                    ko_level=round(init_p * p.ko_pct, 4) if init_p and p.ko_pct else None,
                    strike_level=round(init_p * p.strike_pct, 4) if init_p and p.strike_pct else None,
                    eki_level=round(init_p * p.eki_pct, 4) if init_p and p.eki_pct else None,
                    ko_hit=False, position_order=i,
                )
                db.session.add(u)

        db.session.commit()
        flash(f'{p.product_code} 已更新', 'success')
        if p.status == 'ko_exited':
            return redirect(url_for('ko_history'))
        return redirect(url_for('product_detail', pid=p.id))

    underlyings = sorted(p.underlyings, key=lambda u: u.position_order or 0)
    return render_template('products/edit.html', product=p, underlyings=underlyings)


# ── 部位管理（多客戶同一商品）────────────────────────────────────────────────
@app.route('/products/<int:pid>/add_position', methods=['POST'])
@login_required
def add_position(pid):
    f = request.form
    pos = Position(
        client_id         = int(f['client_id']),
        product_id        = pid,
        investment_amount = float(f['investment_amount']),
        notes             = f.get('notes', ''),
    )
    db.session.add(pos)
    db.session.commit()
    flash('已新增部位', 'success')
    return redirect(url_for('product_detail', pid=pid))


# ── API：新增客戶（AJAX）─────────────────────────────────────────────────────
@app.route('/api/clients', methods=['POST'])
@login_required
def api_add_client():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '請輸入客戶姓名'}), 400
    existing = Client.query.filter_by(name=name, user_id=current_uid()).first()
    if existing:
        return jsonify({'id': existing.id, 'name': existing.name_masked})
    c = Client(name=name, name_masked=Client.mask_name(name), user_id=current_uid())
    db.session.add(c)
    db.session.commit()
    return jsonify({'id': c.id, 'name': c.name_masked})


# ── API：每日持倉摘要（給外部排程系統呼叫，如「客戶日報系統」） ──────────────
@app.route('/api/daily_summary')
def api_daily_summary():
    """以 token 認證，回傳指定 user 的 FCN 持倉摘要 JSON。
    用途：給「客戶日報系統」每日 06:30 抓取，整合進個人 briefing email。

    參數：
      - token    必填，需與 env DAILY_SUMMARY_TOKEN 相符

    回傳的 user 由伺服器 env DAILY_SUMMARY_USERNAME 決定（預設 'admin'）。
    URL 不接受 ?username= 覆寫，避免 token 外洩時被拿去查其他帳號。
    此 endpoint 為「唯讀」，不會修改任何資料。

    回傳：
      {
        "as_of": "YYYY-MM-DD",
        "user": "admin",
        "active_count": int,
        "near_ko_threshold_pct": -3.0,
        "expiring_soon_days": 14,
        "upcoming_payment_days": 7,
        "summary": {
          "near_ko":          [{product_code, ticker, distance_pct, latest_price, ko_level, price_date}, ...],
          "expiring_soon":    [{product_code, days_to_maturity, maturity_date}, ...],
          "upcoming_payments": [{product_code, period, days_to, payment_date, monthly_coupon_per_position}, ...],
          "yesterday_ko":     [{product_code, ticker, price, ko_level, ko_hit_date}, ...],
          "yesterday_matured":[{product_code, maturity_date}, ...],
          "overdue":          [{product_code, days_overdue, maturity_date}, ...]
        }
      }
    """
    import os, json as _json2

    expected_token = os.environ.get('DAILY_SUMMARY_TOKEN', '')
    if not expected_token:
        return jsonify({'error': 'DAILY_SUMMARY_TOKEN 未設定於伺服器環境變數'}), 503
    if request.args.get('token', '') != expected_token:
        return jsonify({'error': 'unauthorized'}), 401

    # 鎖死只能讀 env 指定的帳號（預設 admin），不接受 URL 覆寫
    username = os.environ.get('DAILY_SUMMARY_USERNAME', 'admin')
    user = AppUser.query.filter_by(username=username).first()
    if not user:
        return jsonify({'error': f'伺服器設定的 username 不存在: {username}'}), 404

    # 取「最近一個美股交易日」收盤判定
    from daily_report import _get_last_us_trading_day, _is_ko_relevant_today
    last_td = _get_last_us_trading_day()
    today = date.today()

    active = (Product.query
              .filter_by(status='active', user_id=user.id)
              .order_by(Product.trade_date, Product.created_at)
              .all())

    near_ko, expiring_soon, upcoming_payments = [], [], []
    yesterday_ko, yesterday_matured, overdue = [], [], []

    for p in active:
        code = p.product_code or '?'

        # 接近 KO（≤3% 且尚未鎖定、且今天會比 KO）
        if _is_ko_relevant_today(p):
            for u in sorted(p.underlyings, key=lambda x: x.position_order or 0):
                if not u.latest_price or not u.ko_level or u.ko_hit:
                    continue
                dist = u.latest_price / u.ko_level - 1
                if -0.03 <= dist < 0:
                    near_ko.append({
                        'product_code': code,
                        'ticker': u.ticker,
                        'distance_pct': round(dist * 100, 2),
                        'latest_price': round(u.latest_price, 2),
                        'ko_level': round(u.ko_level, 2),
                        'price_date': u.price_date.isoformat() if u.price_date else None,
                    })

        # 即將到期（1-14 天）
        if p.maturity_date and p.days_to_maturity is not None:
            if 1 <= p.days_to_maturity <= 14:
                expiring_soon.append({
                    'product_code': code,
                    'days_to_maturity': p.days_to_maturity,
                    'maturity_date': p.maturity_date.isoformat(),
                })
            elif p.days_to_maturity < 0:
                overdue.append({
                    'product_code': code,
                    'days_overdue': abs(p.days_to_maturity),
                    'maturity_date': p.maturity_date.isoformat(),
                })

        # 昨日達 KO（任一標的 ko_hit_date == 最近交易日）
        for u in p.underlyings:
            if u.ko_hit and u.ko_hit_date == last_td:
                yesterday_ko.append({
                    'product_code': code,
                    'ticker': u.ticker,
                    'price': round(u.latest_price, 2) if u.latest_price else None,
                    'ko_level': round(u.ko_level, 2) if u.ko_level else None,
                    'ko_hit_date': last_td.isoformat(),
                })

        # 昨日到期
        if p.maturity_date == last_td:
            yesterday_matured.append({
                'product_code': code,
                'maturity_date': p.maturity_date.isoformat(),
            })

        # 即將配息（7 天內，含當天）
        for s in p.payment_schedule:
            if s.payment_date and 0 <= (s.payment_date - today).days <= 7:
                # 取 position 1 美金的月配息（給排程系統當參考）
                monthly_rate = (p.coupon_rate / 12) if p.coupon_rate else None
                upcoming_payments.append({
                    'product_code': code,
                    'period': s.period,
                    'days_to': (s.payment_date - today).days,
                    'payment_date': s.payment_date.isoformat(),
                    'monthly_coupon_per_usd': round(monthly_rate, 6) if monthly_rate else None,
                })

    # 排序
    near_ko.sort(key=lambda x: x['distance_pct'])               # 最接近 KO 在前
    expiring_soon.sort(key=lambda x: x['days_to_maturity'])     # 最快到期在前
    upcoming_payments.sort(key=lambda x: x['days_to'])          # 最近配息在前

    return jsonify({
        'as_of': today.isoformat(),
        'last_trading_day': last_td.isoformat(),
        'user': username,
        'active_count': len(active),
        'near_ko_threshold_pct': -3.0,
        'expiring_soon_days': 14,
        'upcoming_payment_days': 7,
        'summary': {
            'near_ko': near_ko,
            'expiring_soon': expiring_soon,
            'upcoming_payments': upcoming_payments,
            'yesterday_ko': yesterday_ko,
            'yesterday_matured': yesterday_matured,
            'overdue': overdue,
        },
    })


# ── API：FCN 售後事件（C 模組：配息 / 到期 / KO 接股票） ───────────────────
@app.route('/api/aftercare_events')
def api_aftercare_events():
    """以 token 認證，回傳今日該追蹤的 FCN 售後事件 + 涉及客戶名單。

    事件規則：
      - payment: 明天配息 (d=1) 或今天配息 (d=0)
      - maturity: 7 天內到期（天天提醒）
      - ko: 昨日達 KO（接股票流程啟動）

    user 鎖定 DAILY_SUMMARY_USERNAME（預設 admin），URL 不接受覆寫。
    """
    import os
    expected = os.environ.get('DAILY_SUMMARY_TOKEN', '')
    if not expected:
        return jsonify({'error': 'DAILY_SUMMARY_TOKEN 未設定'}), 503
    if request.args.get('token', '') != expected:
        return jsonify({'error': 'unauthorized'}), 401

    username = os.environ.get('DAILY_SUMMARY_USERNAME', 'admin')
    user = AppUser.query.filter_by(username=username).first()
    if not user:
        return jsonify({'error': f'username 不存在: {username}'}), 404

    from daily_report import _get_last_us_trading_day
    last_td = _get_last_us_trading_day()
    today = date.today()

    active = (Product.query
              .filter_by(status='active', user_id=user.id)
              .order_by(Product.trade_date, Product.created_at)
              .all())

    def _client_brief(pos):
        return {
            'client_id': pos.client.id,
            'name_masked': pos.client.name_masked,
            'amount': pos.investment_amount,
            'currency': pos.product.currency or 'USD',
            'tags': pos.client.tags_list,
            'risk_profile': pos.client.risk_profile,
        }

    payment_events, maturity_events, ko_events = [], [], []

    for p in active:
        code = p.product_code or '?'

        # 配息：明天 (1) 或今天 (0)
        for s in p.payment_schedule:
            if not s.payment_date:
                continue
            d = (s.payment_date - today).days
            if d in (0, 1):
                clients = []
                for pos in p.positions:
                    if not pos.client:
                        continue
                    cb = _client_brief(pos)
                    if p.coupon_rate and pos.investment_amount:
                        cb['estimated_coupon'] = round(pos.investment_amount * p.coupon_rate / 12, 2)
                    clients.append(cb)
                payment_events.append({
                    'type': 'payment',
                    'product_code': code,
                    'event_date': s.payment_date.isoformat(),
                    'days_to': d,  # 0=今天, 1=明天
                    'when_label': '今天' if d == 0 else '明天',
                    'period': s.period,
                    'total_periods': p.total_periods,
                    'coupon_rate_annual': p.coupon_rate,
                    'clients': clients,
                })

        # 到期：7 天內，天天提醒
        if p.maturity_date and p.days_to_maturity is not None and 0 <= p.days_to_maturity <= 7:
            clients = [_client_brief(pos) for pos in p.positions if pos.client]
            maturity_events.append({
                'type': 'maturity',
                'product_code': code,
                'event_date': p.maturity_date.isoformat(),
                'days_to_maturity': p.days_to_maturity,
                'when_label': f'{p.days_to_maturity} 天後' if p.days_to_maturity > 0 else '今天',
                'paid_periods': p.paid_count,
                'total_periods': p.total_periods,
                'clients': clients,
            })

        # KO：昨日達 KO
        ko_tickers = []
        for u in p.underlyings:
            if u.ko_hit and u.ko_hit_date == last_td:
                ko_tickers.append({
                    'ticker': u.ticker,
                    'ko_price': round(u.latest_price, 2) if u.latest_price else None,
                    'ko_level': round(u.ko_level, 2) if u.ko_level else None,
                })
        if ko_tickers:
            clients = [_client_brief(pos) for pos in p.positions if pos.client]
            ko_events.append({
                'type': 'ko',
                'product_code': code,
                'event_date': last_td.isoformat(),
                'when_label': '昨日',
                'ko_tickers': ko_tickers,
                'strike_pct': p.strike_pct,
                'clients': clients,
            })

    payment_events.sort(key=lambda x: x['days_to'])
    maturity_events.sort(key=lambda x: x['days_to_maturity'])

    return jsonify({
        'as_of': today.isoformat(),
        'last_trading_day': last_td.isoformat(),
        'user': username,
        'total_events': len(payment_events) + len(maturity_events) + len(ko_events),
        'payment_events': payment_events,
        'maturity_events': maturity_events,
        'ko_events': ko_events,
    })


# ── API：今日該主動聯絡的客戶（B 模組） ──────────────────────────────────────
# 「Ticker → 對應主題」對照表，用於依持倉自動加分
_TICKER_THEMES = {
    # 半導體
    'NVDA': ['半導體', 'AI 科技'],
    'AMD':  ['半導體', 'AI 科技'],
    'AVGO': ['半導體', 'AI 科技'],
    'TSM':  ['半導體'],
    'INTC': ['半導體'],
    'MU':   ['半導體'],
    'AMAT': ['半導體'],
    'ASML': ['半導體'],
    'ARM':  ['半導體', 'AI 科技'],
    'SMCI': ['半導體', 'AI 科技'],
    'QCOM': ['半導體'],
    # AI / 科技龍頭
    'GOOGL': ['AI 科技', '美股成長'],
    'GOOG':  ['AI 科技', '美股成長'],
    'META':  ['AI 科技', '美股成長'],
    'MSFT':  ['AI 科技', '美股成長'],
    'AAPL':  ['AI 科技', '美股成長'],
    'AMZN':  ['AI 科技', '美股成長'],
    'ORCL':  ['AI 科技'],
    'CRM':   ['AI 科技'],
    'HPE':   ['AI 科技'],
    'DELL':  ['AI 科技'],
    # 美股成長
    'TSLA':  ['美股成長'],
    'NFLX':  ['美股成長'],
    # 金融
    'JPM':   ['金融'], 'BAC': ['金融'], 'GS': ['金融'], 'WFC': ['金融'],
    'MA':    ['金融'], 'V':   ['金融'],
    # 醫療生技
    'JNJ':   ['醫療生技'], 'PFE': ['醫療生技'], 'MRK': ['醫療生技'],
    'LLY':   ['醫療生技'], 'UNH': ['醫療生技'],
    # 能源原物料
    'XOM':   ['能源原物料'], 'CVX': ['能源原物料'], 'COP': ['能源原物料'],
}


@app.route('/api/contact_suggestions')
def api_contact_suggestions():
    """以 token 認證，依今日主題比對 admin 客戶，回傳前 N 位推薦聯絡客戶。

    參數：
      - token  必填，需與 env DAILY_SUMMARY_TOKEN 相符
      - themes 必填，逗號分隔主題清單，如 "半導體,收息"
      - limit  可選，預設 5

    回傳每位客戶含：name_masked、risk_profile、tags、持倉摘要、score、match_reasons。
    user 由 DAILY_SUMMARY_USERNAME 決定，URL 不接受覆寫。
    """
    import os, json as _json4

    expected = os.environ.get('DAILY_SUMMARY_TOKEN', '')
    if not expected:
        return jsonify({'error': 'DAILY_SUMMARY_TOKEN 未設定'}), 503
    if request.args.get('token', '') != expected:
        return jsonify({'error': 'unauthorized'}), 401

    themes_str = request.args.get('themes', '').strip()
    themes = [t.strip() for t in themes_str.split(',') if t.strip()]
    if not themes:
        return jsonify({'error': '請帶 themes 參數，如 ?themes=半導體,收息'}), 400

    limit = min(int(request.args.get('limit', 5)), 20)
    username = os.environ.get('DAILY_SUMMARY_USERNAME', 'admin')
    user = AppUser.query.filter_by(username=username).first()
    if not user:
        return jsonify({'error': f'username 不存在: {username}'}), 404

    # 風險屬性 → 主題傾向（軟性加分）
    rp_theme_affinity = {
        '積極': {'半導體', 'AI 科技', '美股成長'},
        '穩健': {'美股成長', '高股息', '金融', '醫療生技'},
        '保守': {'收息固收', '債券', '高股息'},
    }

    today = date.today()
    suggestions = []

    for c in Client.query.filter_by(user_id=user.id).all():
        score = 0
        reasons = []

        # Tag 直接命中
        client_tags = set(c.tags_list)
        for theme in themes:
            if theme in client_tags:
                score += 3
                reasons.append(f'tag:{theme}')

        # 持倉 ticker → 主題 → 命中
        active_positions = [p for p in c.positions if p.product.status == 'active']
        position_tickers = set()
        for pos in active_positions:
            for u in pos.product.underlyings:
                position_tickers.add(u.ticker.upper())

        for ticker in position_tickers:
            ticker_themes = set(_TICKER_THEMES.get(ticker, []))
            hit = ticker_themes & set(themes)
            if hit:
                score += 2
                for theme in hit:
                    reasons.append(f'持有 {ticker}（{theme}）')

        # 即將配息 / 即將到期（時效性加分）
        for pos in active_positions:
            p = pos.product
            if p.maturity_date:
                d = (p.maturity_date - today).days
                if 0 <= d <= 14:
                    score += 2
                    reasons.append(f'[{p.product_code}] {d}天後到期')
            for s in p.payment_schedule:
                if s.payment_date and 0 <= (s.payment_date - today).days <= 7:
                    score += 1
                    reasons.append(f'[{p.product_code}] {(s.payment_date - today).days}天後配息')
                    break

        # 風險屬性 affinity
        if c.risk_profile and c.risk_profile in rp_theme_affinity:
            if any(t in rp_theme_affinity[c.risk_profile] for t in themes):
                score += 1
                reasons.append(f'風險屬性 {c.risk_profile} 對應主題')

        if score <= 0:
            continue

        suggestions.append({
            'client_id': c.id,
            'client_name_masked': c.name_masked,
            'risk_profile': c.risk_profile,
            'tags': c.tags_list,
            'score': score,
            'match_reasons': reasons,
            'positions_summary': [
                {
                    'product_code': pos.product.product_code,
                    'tickers': [u.ticker for u in pos.product.underlyings],
                    'amount': pos.investment_amount,
                    'currency': pos.product.currency,
                }
                for pos in active_positions
            ],
            'notes': c.notes,
        })

    # 排序 + 取前 N
    suggestions.sort(key=lambda x: x['score'], reverse=True)
    suggestions = suggestions[:limit]

    return jsonify({
        'as_of': today.isoformat(),
        'user': username,
        'themes_received': themes,
        'total_matched': len(suggestions),
        'suggestions': suggestions,
    })


# ── 每日晨會快報 ─────────────────────────────────────────────────────────────
@app.route('/briefing')
@login_required
def briefing():
    from datetime import timedelta
    today = date.today()
    uid = current_uid()

    active = Product.query.filter_by(status='active', user_id=uid).order_by(Product.trade_date, Product.created_at).all()

    # 即將到期（30天內）
    expiring = [p for p in active if p.maturity_date and p.days_to_maturity is not None and 0 <= p.days_to_maturity <= 30]

    # 標的異動（漲跌超過3%）+ KO/Strike/EKI 警示
    alerts = []
    for p in active:
        for u in sorted(p.underlyings, key=lambda x: x.position_order or 0):
            if not u.latest_price or not u.initial_price:
                continue
            change_pct = (u.latest_price - u.initial_price) / u.initial_price
            alert = {
                'product_code': p.product_code,
                'ticker': u.ticker,
                'ticker_name': TICKER_NAME.get(u.ticker, ''),
                'latest_price': u.latest_price,
                'price_date': u.price_date,
                'change_pct': change_pct,
                'ko_dist': (u.latest_price / u.ko_level - 1) if u.ko_level else None,
                'strike_dist': (u.latest_price / u.strike_level - 1) if u.strike_level else None,
                'eki_dist': (u.latest_price / u.eki_level - 1) if u.eki_level else None,
                'ko_hit': u.ko_hit,
                'hit_strike': u.strike_level and u.latest_price <= u.strike_level,
                'hit_eki': u.eki_level and u.latest_price <= u.eki_level,
            }
            alerts.append(alert)

    # 市場指標（多來源備援）
    market = None
    try:
        from price_fetcher import fetch_indices
        market = fetch_indices()
    except Exception:
        market = None

    # Fear & Greed Index
    fear_greed = None
    try:
        import requests as _req
        r = _req.get('https://production.dataviz.cnn.io/index/fearandgreed/current',
                     headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'},
                     timeout=5)
        if r.ok:
            data = r.json()
            fear_greed = {'score': round(data.get('score', 0)), 'rating': data.get('rating', '')}
    except:
        fear_greed = None

    # VIX 指數
    vix = None
    try:
        import yfinance as yf
        vix_data = yf.Ticker('^VIX').history(period='2d')
        if not vix_data.empty:
            last = float(vix_data['Close'].iloc[-1])
            prev = float(vix_data['Close'].iloc[-2]) if len(vix_data) >= 2 else last
            vix_chg = round((last / prev - 1) * 100, 2) if prev else 0
            vix = {'price': round(last, 2), 'chg': vix_chg}
    except:
        vix = None

    return render_template('briefing.html', today=today, active=active,
                           expiring=expiring, alerts=alerts, market=market,
                           fear_greed=fear_greed, vix=vix)


# ── 客戶 PDF 報告 ────────────────────────────────────────────────────────────

def _make_single_product_pdf(pos, today, font_path):
    """產生單一商品的 PDF，回傳 BytesIO"""
    from fpdf import FPDF
    from io import BytesIO
    import tempfile
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    import matplotlib.font_manager as fm
    # 設定中文字型
    _zh_font = None
    _font_candidates = [
        r'C:\Windows\Fonts\msjh.ttc',
        r'C:\Windows\Fonts\msyh.ttc',
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts', 'NotoSansTC-Static.ttf'),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts', 'NotoSansTC-Variable.ttf'),
    ]
    for fp in _font_candidates:
        if os.path.exists(fp):
            _zh_font = fm.FontProperties(fname=fp)
            fm.fontManager.addfont(fp)
            plt.rcParams['font.family'] = _zh_font.get_name()
            break
    if not _zh_font:
        plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei', 'Noto Sans TC', 'SimHei']
    plt.rcParams['axes.unicode_minus'] = False

    p = pos.product

    class PDF(FPDF):
        def header(self):
            if self.page_no() > 1:
                self.set_font('chinese', '', 8)
                self.cell(0, 5, '追蹤報告', align='R')
                self.ln(8)
        def footer(self):
            self.set_y(-15)
            self.set_font('chinese', '', 8)
            self.set_text_color(150, 150, 150)
            self.cell(0, 10, f'第 {self.page_no()} 頁', align='C')

    pdf = PDF(orientation='L')
    if font_path and os.path.exists(font_path):
        pdf.add_font('chinese', '', font_path, uni=True)
    else:
        pdf.add_font('chinese', '', uni=True)
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # 標題
    pdf.set_font('chinese', '', 28)
    pdf.cell(0, 16, '追蹤報告', align='C', new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('chinese', '', 14)
    pdf.cell(0, 10, f'報告日期：{today.strftime("%Y/%m/%d")}', align='C', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(4)

    # 摘要
    pdf.set_font('chinese', '', 13)
    pdf.set_fill_color(240, 240, 240)
    _cur = p.currency or 'USD'
    _amt = f'{pos.investment_amount:,.0f}' if pos.investment_amount else '-'
    _coupon = f'{pos.monthly_coupon:,.0f}' if pos.monthly_coupon else '-'
    pdf.cell(0, 10, f'投資金額：{_cur} {_amt}    每月配息：{_cur} {_coupon}',
             align='C', fill=True, new_x='LMARGIN', new_y='NEXT')
    pdf.ln(5)

    # 先並行抓取所有標的歷史資料（大幅加速）
    from concurrent.futures import ThreadPoolExecutor
    from price_fetcher import fetch_history

    uls = sorted(p.underlyings, key=lambda u: u.position_order or 0)
    start_d = today - timedelta(days=365)
    history_cache = {}

    def _fetch_one(ticker):
        try:
            return ticker, fetch_history(ticker, start_d, today)
        except Exception:
            return ticker, None

    tickers_to_fetch = [u.ticker for u in uls if u.ticker]
    with ThreadPoolExecutor(max_workers=len(tickers_to_fetch) or 1) as pool:
        for ticker, data in pool.map(_fetch_one, tickers_to_fetch):
            if data is not None and not data.empty:
                history_cache[ticker] = data

    # 圖表產生函數
    chart_files = []

    def make_chart(ticker_symbol, product_code, ko_level, strike_level, eki_level, strike_pct=None, eki_pct=None, ticker_name=''):
        data = history_cache.get(ticker_symbol) if 'history_cache' in dir() or 'history_cache' in locals() else None
        if data is None:
            try:
                from price_fetcher import fetch_history
                from datetime import timedelta
                start_d = today - timedelta(days=365)
                data = fetch_history(ticker_symbol, start_d, today)
                if data is None or data.empty:
                    return None
            except Exception as e:
                app.logger.warning(f'Chart data fetch failed for {ticker_symbol}: {e}')
                return None
        try:
            close = data['Close'].squeeze()
            volume = data['Volume'].squeeze() if 'Volume' in data.columns else close * 0
            fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8, 4.5), height_ratios=[3, 1], sharex=True)
            fig.subplots_adjust(hspace=0.05)
            _display = f'{ticker_symbol} {ticker_name}' if ticker_name else ticker_symbol
            ax1.plot(close.index, close.values, color='#e74c3c', linewidth=2.5, label=_display)
            if ko_level:
                ax1.axhline(y=ko_level, color='#e74c3c', linestyle='--', linewidth=1.5, label=f'期初價格({ko_level:,.4f})')
            if strike_level:
                strike_pct_label = f'{strike_pct:.2%}' if strike_pct else ''
                ax1.axhline(y=strike_level, color='#e67e22', linestyle='--', linewidth=1.5, label=f'執行價{strike_pct_label}({strike_level:,.4f})')
            if eki_level:
                eki_pct_label = f'{eki_pct:.0%}' if eki_pct else ''
                ax1.axhline(y=eki_level, color='#8e44ad', linestyle='--', linewidth=1.5, label=f'觸及生效價{eki_pct_label}({eki_level:,.4f})')
            last_price = close.iloc[-1]
            ax1.annotate(f'{last_price:.2f}', xy=(close.index[-1], last_price),
                        fontsize=10, fontweight='bold', color='#e74c3c',
                        xytext=(10, 5), textcoords='offset points')
            _title = f'{ticker_symbol} {ticker_name} - {product_code}' if ticker_name else f'{ticker_symbol} - {product_code}'
            ax1.set_title(_title, fontsize=13, fontweight='bold', pad=10)
            ax1.legend(loc='upper left', fontsize=8, framealpha=0.9)
            ax1.set_ylabel('Price (USD)', fontsize=9)
            ax1.grid(True, alpha=0.3)
            ax1.spines['top'].set_visible(False)
            ax1.spines['right'].set_visible(False)
            colors = ['#e74c3c' if i > 0 and close.iloc[i] < close.iloc[i-1] else '#27ae60' for i in range(len(close))]
            ax2.bar(volume.index, volume.values, color=colors, alpha=0.7, width=1)
            ax2.set_ylabel('Volume', fontsize=8)
            ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y/%m'))
            ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
            ax2.grid(True, alpha=0.3)
            ax2.spines['top'].set_visible(False)
            ax2.spines['right'].set_visible(False)
            plt.tight_layout()
            tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            plt.savefig(tmp.name, dpi=150)
            plt.close(fig)
            return tmp.name
        except Exception as e:
            app.logger.warning(f'Chart render failed: {e}')
            return None

    # 持倉明細
    has_eki = any(u.eki_level for u in uls)

    # 商品標題
    pdf.set_font('chinese', '', 13)
    pdf.set_fill_color(26, 26, 46)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 9, f'  {p.product_code}  |  {p.issuer or "-"}  |  {p.tenor_months}M  |  {"{:.1%}".format(p.coupon_rate) if p.coupon_rate else "-"}',
             fill=True, new_x='LMARGIN', new_y='NEXT')
    pdf.set_text_color(0, 0, 0)

    # 基本資訊
    pdf.set_font('chinese', '', 10)
    days = p.days_to_maturity
    days_str = f'{days} 天' if days is not None else '-'
    trade_date_str = p.trade_date.strftime('%Y/%m/%d') if p.trade_date else '-'
    start_date_str = p.start_date.strftime('%Y/%m/%d') if p.start_date else '-'
    maturity_date_str = p.maturity_date.strftime('%Y/%m/%d') if p.maturity_date else '-'
    _cur2 = p.currency or 'USD'
    _amt2 = f'{pos.investment_amount:,.0f}' if pos.investment_amount else '-'
    _coupon2 = f'{pos.monthly_coupon:,.0f}' if pos.monthly_coupon else '-'
    pdf.cell(0, 7, f'投資金額：{_cur2} {_amt2}    月配息：{_cur2} {_coupon2}    交易日：{trade_date_str}    比價日：{start_date_str}    到期日：{maturity_date_str}    剩餘：{days_str}',
             new_x='LMARGIN', new_y='NEXT')

    # 標的表格
    pdf.set_font('chinese', '', 9)
    pdf.set_fill_color(220, 220, 220)
    strike_pct_str = f'({p.strike_pct:.2%})' if p.strike_pct else ''
    # 最新價日期
    price_date_str = ''
    for u in uls:
        if u.price_date:
            price_date_str = f'({u.price_date.strftime("%m/%d")})'
            break
    if has_eki:
        col_w = [50, 32, 36, 36, 36, 32, 25]
        eki_pct_str = f'({p.eki_pct:.0%})' if p.eki_pct else ''
        headers = ['連結標的', '期初價格', '提前出場價', f'執行價{strike_pct_str}', f'觸及生效價{eki_pct_str}', f'最新價{price_date_str}', '記憶式出場']
    else:
        col_w = [55, 38, 42, 42, 38, 25]
        headers = ['連結標的', '期初價格', '提前出場價', f'執行價{strike_pct_str}', f'最新價{price_date_str}', '記憶式出場']

    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1, fill=True, align='C')
    pdf.ln()

    for u in uls:
        ticker_cn = TICKER_NAME.get(u.ticker, '')
        ticker_display = f'{u.ticker} {ticker_cn}' if ticker_cn else u.ticker
        if has_eki:
            vals = [
                ticker_display,
                f'{u.initial_price:,.4f}' if u.initial_price else '-',
                f'{u.ko_level:,.4f}' if u.ko_level else '-',
                f'{u.strike_level:,.4f}' if u.strike_level else '-',
                f'{u.eki_level:,.4f}' if u.eki_level else '-',
                f'{u.latest_price:,.2f}' if u.latest_price else '-',
                'V' if u.ko_hit else '',
            ]
        else:
            vals = [
                ticker_display,
                f'{u.initial_price:,.4f}' if u.initial_price else '-',
                f'{u.ko_level:,.4f}' if u.ko_level else '-',
                f'{u.strike_level:,.4f}' if u.strike_level else '-',
                f'{u.latest_price:,.2f}' if u.latest_price else '-',
                'V' if u.ko_hit else '',
            ]
        for i, v in enumerate(vals):
            pdf.cell(col_w[i], 6.5, v, border=1, align='C')
        pdf.ln()

    pdf.ln(2)

    # ── 配息排程表（第一頁紅框區）──
    sched = sorted(p.payment_schedule, key=lambda s: s.period) if p.payment_schedule else []
    if sched:
        np_obj = p.next_payment
        pdf.set_font('chinese', '', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 8, f'配息排程（已配 {p.paid_count} / 共 {len(sched)} 期）',
                 new_x='LMARGIN', new_y='NEXT')
        cw = [16, 44, 44, 28]          # 期數 / 觀察期間結束日 / 配息日 / 狀態
        grp_w = sum(cw)
        gap = 10
        sub_headers = ['期數', '觀察期間結束日', '配息日', '狀態']
        x0 = pdf.l_margin
        y0 = pdf.get_y()
        # 表頭（左右兩組）
        pdf.set_font('chinese', '', 8)
        for gi in range(2):
            pdf.set_xy(x0 + gi * (grp_w + gap), y0)
            pdf.set_fill_color(210, 210, 210)
            pdf.set_text_color(0, 0, 0)
            for i, h in enumerate(sub_headers):
                pdf.cell(cw[i], 6, h, border=1, fill=True, align='C')
        # 資料列：分左右兩欄
        n = len(sched)
        half = (n + 1) // 2
        groups = [sched[:half], sched[half:]]
        for r in range(half):
            for gi, grp in enumerate(groups):
                if r >= len(grp):
                    continue
                s = grp[r]
                pdf.set_xy(x0 + gi * (grp_w + gap), y0 + 6 + r * 6)
                if s.payment_date and s.payment_date < today:
                    pdf.set_fill_color(235, 235, 235)
                    pdf.set_text_color(135, 135, 135)
                    st_txt = '已配息'
                elif np_obj and s.id == np_obj.id:
                    pdf.set_fill_color(13, 110, 253)
                    pdf.set_text_color(255, 255, 255)
                    st_txt = '下次配息'
                else:
                    pdf.set_fill_color(255, 255, 255)
                    pdf.set_text_color(0, 0, 0)
                    st_txt = '未到'
                obs_end = s.obs_end_date.strftime('%Y/%m/%d') if s.obs_end_date else '-'
                pay = s.payment_date.strftime('%Y/%m/%d') if s.payment_date else '-'
                pdf.cell(cw[0], 6, str(s.period), border=1, fill=True, align='C')
                pdf.cell(cw[1], 6, obs_end, border=1, fill=True, align='C')
                pdf.cell(cw[2], 6, pay, border=1, fill=True, align='C')
                pdf.cell(cw[3], 6, st_txt, border=1, fill=True, align='C')
        pdf.set_text_color(0, 0, 0)
        pdf.set_y(y0 + 6 + half * 6 + 4)

    # 線圖（兩張並排）
    chart_paths = []
    for u in uls:
        if not u.ticker:
            continue
        # Stepdown FCN 用期初價格畫「期初價格」線，一般 FCN 用 ko_level
        chart_ko = u.initial_price if p.ko_type == 'stepdown' and u.initial_price else u.ko_level
        chart_path = make_chart(u.ticker, p.product_code, chart_ko, u.strike_level, u.eki_level, p.strike_pct, p.eki_pct, TICKER_NAME.get(u.ticker, ''))
        if chart_path:
            chart_files.append(chart_path)
            chart_paths.append(chart_path)

    if chart_paths:
        for chart_path in chart_paths:
            pdf.add_page('L')
            pdf.image(chart_path, x=15, y=15, w=267)

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)

    # 清理暫存圖檔
    for f in chart_files:
        try:
            os.unlink(f)
        except:
            pass

    return buf


@app.route('/report/<int:cid>/<int:pid>')
@login_required
def client_report_single(cid, pid):
    """單一商品 PDF 報告"""
    log_activity(f'客戶報告(單一) cid={cid} pid={pid}')
    from flask import send_file
    from setup_fonts import get_font_paths

    client = Client.query.get_or_404(cid)
    pos = Position.query.filter_by(client_id=cid, product_id=pid).first_or_404()
    if pos.product.user_id != current_uid():
        abort(403)

    today = date.today()
    font_path, _ = get_font_paths()
    buf = _make_single_product_pdf(pos, today, font_path)
    fname = f'{pos.product.product_code}_{today.strftime("%Y%m%d")}.pdf'
    return send_file(buf, download_name=fname, as_attachment=True, mimetype='application/pdf')


@app.route('/report/<int:cid>')
@login_required
def client_report(cid):
    """全部商品打包 — 單一商品直接回傳 PDF，多商品回傳 ZIP"""
    log_activity(f'客戶報告(全部) cid={cid}')
    from io import BytesIO
    from flask import send_file
    from setup_fonts import get_font_paths
    import zipfile

    client = Client.query.get_or_404(cid)
    positions = [pos for pos in client.positions if pos.product.user_id == current_uid()]
    active_pos = [pos for pos in positions if pos.product.status == 'active']
    today = date.today()
    font_path, _ = get_font_paths()

    if not active_pos:
        flash('此客戶無有效持倉', 'warning')
        return redirect(url_for('reports'))

    # 單一商品 → 直接回傳 PDF
    if len(active_pos) == 1:
        buf = _make_single_product_pdf(active_pos[0], today, font_path)
        fname = f'{active_pos[0].product.product_code}_{today.strftime("%Y%m%d")}.pdf'
        return send_file(buf, download_name=fname, as_attachment=True, mimetype='application/pdf')

    # 多商品 → 各自產出 PDF 後打包 ZIP
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for pos in active_pos:
            pdf_buf = _make_single_product_pdf(pos, today, font_path)
            fname = f'{pos.product.product_code}_{today.strftime("%Y%m%d")}.pdf'
            zf.writestr(fname, pdf_buf.read())

    zip_buf.seek(0)
    return send_file(zip_buf,
                     download_name=f'{client.name_masked}_報告_{today.strftime("%Y%m%d")}.zip',
                     as_attachment=True, mimetype='application/zip')


# ── 客戶報告列表 ─────────────────────────────────────────────────────────────
@app.route('/reports')
@login_required
def reports():
    all_clients = Client.query.filter_by(user_id=current_uid()).order_by(Client.name).all()
    return render_template('reports.html', clients=all_clients)


# ── 診斷（部署後可刪除）─────────────────────────────────────────────────────
@app.route('/debug_db')
@login_required
def debug_db():
    db_uri = app.config.get('SQLALCHEMY_DATABASE_URI', 'NOT SET')
    import re
    safe_uri = re.sub(r'://[^@]+@', '://***@', db_uri) if db_uri else 'NOT SET'
    uid = current_uid()
    # 跟 dashboard 完全一樣的查詢
    active = Product.query.filter_by(status='active', user_id=uid).order_by(Product.created_at).all()
    products_info = [{'id': p.id, 'code': p.product_code, 'user_id': p.user_id, 'status': p.status, 'created_at': str(p.created_at)} for p in active]
    # 也查不帶 order_by 的
    active2 = Product.query.filter_by(status='active', user_id=uid).all()
    return jsonify({
        'db': safe_uri,
        'session_user_id': session.get('user_id'),
        'current_uid': uid,
        'with_order': len(active),
        'without_order': len(active2),
        'products': products_info,
    })


# ── 工具函數 ──────────────────────────────────────────────────────────────────
def _parse_date(s):
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except:
        return None


# ── 上傳 TS → 客戶文件圖片 ──────────────────────────────────────────────────
@app.route('/upload_ts', methods=['GET', 'POST'])
@login_required
def upload_ts():
    if request.method == 'POST':
        log_activity('上傳 TS')
        from client_doc_generator import generate_client_doc_image
        from flask import send_file
        import tempfile

        f = request.files.get('ts_file')
        if not f or not f.filename.endswith('.pdf'):
            flash('請上傳 PDF 檔案', 'danger')
            return redirect(url_for('upload_ts'))

        # 存暫存檔
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        f.save(tmp.name)
        tmp.close()

        try:
            buf, data = generate_client_doc_image(tmp.name)
            product_code = data.get('product_code', 'unknown')
            filename = f'{product_code}_客戶文件.png'
            return send_file(buf, download_name=filename, as_attachment=True, mimetype='image/png')
        except Exception as e:
            flash(f'解析失敗：{str(e)}', 'danger')
            return redirect(url_for('upload_ts'))
        finally:
            os.unlink(tmp.name)

    return render_template('upload_ts.html')


# ── 報價圖片產生器 ─────────────────────────────────────────────────────────────
@app.route('/quote')
@login_required
def quote():
    return render_template('quote.html')


@app.route('/quote/generate', methods=['POST'])
@login_required
def generate_quote():
    log_activity('報價產生')
    from quote_generator import generate_quote_image
    from flask import send_file

    from quote_generator import TICKER_NAME as QT_TICKER_NAME
    tickers = request.form.getlist('tickers')

    # 處理自訂標的
    custom_text = request.form.get('custom_tickers', '').strip()
    if custom_text:
        for line in custom_text.splitlines():
            line = line.strip()
            if not line:
                continue
            parts = line.split(None, 1)
            if parts:
                ticker = parts[0].upper()
                if ticker not in tickers:
                    tickers.append(ticker)
                if len(parts) > 1:
                    QT_TICKER_NAME[ticker] = parts[1]

    if not tickers:
        flash('請至少選擇一個連結標的', 'danger')
        return redirect(url_for('quote'))

    # 解析百分比（使用者輸入 72.05 → 0.7205）
    def parse_pct(val):
        if not val:
            return None
        try:
            v = float(val)
            if v > 1:
                return v / 100
            return v
        except:
            return None

    params = {
        'product_type': request.form.get('product_type', 'FCN'),
        'tenor': request.form.get('tenor', '6M'),
        'strike_pct': parse_pct(request.form.get('strike_pct')),
        'ko_pct': parse_pct(request.form.get('ko_pct')),
        'coupon': parse_pct(request.form.get('coupon')),
        'eki_pct': parse_pct(request.form.get('eki_pct')),
        'ko_start': request.form.get('ko_start', '1M'),
        'memory_ko': request.form.get('memory_ko') == '1',
        'currency': request.form.get('currency', 'USD'),
        'issuer': request.form.get('issuer', ''),
        'tickers': tickers,
        'ko_start_pct_q': parse_pct(request.form.get('ko_start_pct_q')),
        'ko_stepdown_q': parse_pct(request.form.get('ko_stepdown_q')),
    }
    # Stepdown FCN: 用起始KO算 ko_pct 和排程
    if params['product_type'] == 'Stepdown FCN' and params['ko_start_pct_q'] and params['ko_stepdown_q']:
        tenor_n = int(params['tenor'].replace('M', ''))
        start_ko = params['ko_start_pct_q']
        step = params['ko_stepdown_q']
        params['ko_schedule'] = [start_ko - step * i for i in range(tenor_n)]
        params['ko_pct'] = params['ko_schedule'][0]  # 第一個月的 KO 當提前出場價

    if not params['strike_pct'] or not params['coupon']:
        flash('Strike 和 Coupon 為必填', 'danger')
        return redirect(url_for('quote'))

    try:
        buf, price_date = generate_quote_image(params)

        # 如果選擇 PNG，直接回傳圖片
        output_format = request.form.get('output_format', 'png')
        if output_format == 'png':
            filename = f'報價_{date.today().strftime("%Y%m%d")}.png'
            return send_file(buf, download_name=filename, as_attachment=True, mimetype='image/png')

        # 產出 PDF：第一頁報價圖 + 後面每個標的線圖
        from fpdf import FPDF
        from io import BytesIO
        import tempfile
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        import matplotlib.font_manager as fm

        # 中文字型
        _font_found = False
        for fp in [r'C:\Windows\Fonts\msjh.ttc', r'C:\Windows\Fonts\msyh.ttc',
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts', 'NotoSansTC-Static.ttf'),
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts', 'NotoSansTC-Variable.ttf')]:
            if os.path.exists(fp):
                fm.fontManager.addfont(fp)
                plt.rcParams['font.family'] = fm.FontProperties(fname=fp).get_name()
                _font_found = True
                break
        if not _font_found:
            plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei', 'Noto Sans TC', 'SimHei']
        plt.rcParams['axes.unicode_minus'] = False

        # 報價圖存暫存檔
        quote_tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
        quote_tmp.write(buf.getvalue())
        quote_tmp.close()

        pdf = FPDF(orientation='L', unit='mm', format='A4')
        font_path = r'C:\Windows\Fonts\msjh.ttc'
        if os.path.exists(font_path):
            pdf.add_font('chinese', '', font_path, uni=True)
        pdf.set_auto_page_break(auto=True, margin=10)

        # 第一頁：報價圖片
        pdf.add_page()
        from PIL import Image as PILImage
        img = PILImage.open(quote_tmp.name)
        img_w, img_h = img.size
        page_w, page_h = 297, 210  # A4 橫向 mm
        ratio = min((page_w - 20) / (img_w * 0.264583), (page_h - 20) / (img_h * 0.264583))
        w_mm = img_w * 0.264583 * ratio
        h_mm = img_h * 0.264583 * ratio
        x = (page_w - w_mm) / 2
        y = (page_h - h_mm) / 2
        pdf.image(quote_tmp.name, x=x, y=y, w=w_mm)

        # 線圖頁面
        strike_pct = params.get('strike_pct')
        eki_pct = params.get('eki_pct')
        ko_pct = params.get('ko_pct')
        chart_files = []

        from quote_generator import fetch_closing_prices, TICKER_NAME as QT_TN
        prices_data, _ = fetch_closing_prices(tickers)

        # 多執行緒同時抓歷史資料 + 畫圖（大幅加速）
        from concurrent.futures import ThreadPoolExecutor
        from price_fetcher import fetch_history
        from datetime import timedelta
        start_d = date.today() - timedelta(days=365)

        # 先並行抓所有歷史資料
        def _fetch(ticker):
            try:
                return ticker, fetch_history(ticker, start_d, date.today())
            except:
                return ticker, None

        history_map = {}
        with ThreadPoolExecutor(max_workers=len(tickers)) as pool:
            for ticker, data in pool.map(lambda t: _fetch(t), tickers):
                if data is not None and not data.empty:
                    history_map[ticker] = data

        # 再逐一畫圖（matplotlib 不支援多執行緒）
        for ticker in tickers:
            data = history_map.get(ticker)
            if data is None:
                continue
            try:
                close = data['Close'].squeeze()
                volume = data['Volume'].squeeze() if 'Volume' in data.columns else close * 0
                latest = prices_data.get(ticker)
                ref_price = latest  # 參考進場價 = 最新收盤價
                strike_level_calc = ref_price * strike_pct if ref_price and strike_pct else None
                eki_level_calc = ref_price * eki_pct if ref_price and eki_pct else None

                fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 5), height_ratios=[3, 1], sharex=True)
                fig.subplots_adjust(hspace=0.05)

                ticker_cn = QT_TN.get(ticker, TICKER_NAME.get(ticker, ''))
                ax1.plot(close.index, close.values, color='#e74c3c', linewidth=2.5, label=f'{ticker} {ticker_cn}')

                if ref_price:
                    ax1.axhline(y=ref_price, color='#e74c3c', linestyle='--', linewidth=1.5,
                               label=f'參考進場價({ref_price:,.2f})')
                if strike_level_calc:
                    ax1.axhline(y=strike_level_calc, color='#e67e22', linestyle='--', linewidth=1.5,
                               label=f'執行價{strike_pct:.0%}({strike_level_calc:,.2f})')
                if eki_level_calc:
                    ax1.axhline(y=eki_level_calc, color='#8e44ad', linestyle='--', linewidth=1.5,
                               label=f'觸及生效價{eki_pct:.0%}({eki_level_calc:,.2f})')

                last_price = close.iloc[-1]
                ax1.annotate(f'{last_price:.2f}', xy=(close.index[-1], last_price),
                            fontsize=10, fontweight='bold', color='#e74c3c',
                            xytext=(10, 5), textcoords='offset points')

                ax1.set_title(f'{ticker} {ticker_cn}', fontsize=14, fontweight='bold', pad=10)
                ax1.legend(loc='upper left', fontsize=9, framealpha=0.9)
                ax1.set_ylabel('Price (USD)', fontsize=9)
                ax1.grid(True, alpha=0.3)
                ax1.spines['top'].set_visible(False)
                ax1.spines['right'].set_visible(False)

                colors_bar = ['#e74c3c' if i > 0 and close.iloc[i] < close.iloc[i-1] else '#27ae60' for i in range(len(close))]
                ax2.bar(volume.index, volume.values, color=colors_bar, alpha=0.7, width=1)
                ax2.set_ylabel('Volume', fontsize=8)
                ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y/%m'))
                ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
                ax2.grid(True, alpha=0.3)
                ax2.spines['top'].set_visible(False)
                ax2.spines['right'].set_visible(False)
                plt.tight_layout()

                tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                plt.savefig(tmp.name, dpi=150)
                plt.close(fig)
                chart_files.append(tmp.name)
            except Exception as e:
                app.logger.warning(f'Quote chart failed for {ticker}: {e}')
                continue

        # 線圖一頁一張
        if chart_files:
            for chart_path in chart_files:
                pdf.add_page()
                pdf.image(chart_path, x=15, y=15, w=267)

        pdf_buf = BytesIO()
        pdf.output(pdf_buf)
        pdf_buf.seek(0)

        # 清理暫存
        try:
            os.unlink(quote_tmp.name)
        except:
            pass
        for f in chart_files:
            try:
                os.unlink(f)
            except:
                pass

        filename = f'{date.today().strftime("%Y-%m-%d")}參考報價.pdf'
        return send_file(pdf_buf, download_name=filename, as_attachment=True, mimetype='application/pdf')
    except Exception as e:
        flash(f'產生失敗：{str(e)}', 'danger')
        return redirect(url_for('quote'))


# ── LINE Webhook ─────────────────────────────────────────────────────────────
@app.route('/line/webhook', methods=['POST'])
def line_webhook():
    from line_bot import get_handler
    handler = get_handler()
    if not handler:
        return 'LINE Bot not configured', 503
    signature = request.headers.get('X-Line-Signature', '')
    body = request.get_data(as_text=True)
    try:
        handler.handle(body, signature)
    except Exception as e:
        app.logger.error(f'LINE Webhook error: {e}')
        return 'Error', 400
    return 'OK', 200


# ── 操作紀錄 ────────────────────────────────────────────────────────────────
@app.route('/activity')
@login_required
def activity_log():
    if not session.get('is_admin'):
        flash('僅管理員可查看', 'danger')
        return redirect(url_for('dashboard'))
    logs = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(200).all()
    return render_template('activity.html', logs=logs)


# ── 市場日報（網頁預覽 + LINE 推播）─────────────────────────────────────────
@app.route('/daily_report')
@login_required
def daily_report():
    from daily_report import generate_market_report, generate_settlement_alert
    from line_bot import is_configured
    market = generate_market_report(app)
    alert = generate_settlement_alert(app)
    if alert:
        report = market + '\n\n' + '─' * 30 + '\n\n' + alert
    else:
        report = market + '\n\n' + '─' * 30 + '\n\n（今日無商品異動，明早不會發送異動提醒）'
    return render_template('daily_report.html', report=report, configured=is_configured())


@app.route('/daily_report/send', methods=['POST'])
@login_required
def send_daily_report():
    from daily_report import generate_market_report, generate_settlement_alert
    from line_bot import broadcast_text, is_configured, get_subscriber_count
    if not is_configured():
        flash('LINE Bot 尚未設定，請先設定 Channel Secret 和 Access Token', 'danger')
        return redirect(url_for('line_settings'))
    # 訊息 A：市場日報
    market = generate_market_report(app)
    success = broadcast_text(market)
    # 訊息 B：商品異動提醒（有事才發）
    alert = generate_settlement_alert(app)
    alert_sent = False
    if alert:
        alert_sent = bool(broadcast_text(alert))
    if success:
        count = get_subscriber_count(app)
        msg = f'市場日報已發送（訂閱人數：{count}）'
        if alert_sent:
            msg += '；商品異動提醒已一併發送'
        elif not alert:
            msg += '；今日無商品異動'
        flash(msg, 'success')
    else:
        flash('LINE 推播失敗，請檢查設定', 'danger')
    return redirect(url_for('daily_report'))


# ── LINE 設定頁 ──────────────────────────────────────────────────────────────
@app.route('/line_settings', methods=['GET', 'POST'])
@login_required
@admin_required
def line_settings():
    from line_bot import is_configured, get_subscriber_count, get_real_follower_count
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if request.method == 'POST':
        secret = request.form.get('channel_secret', '').strip()
        token = request.form.get('access_token', '').strip()
        if secret and token:
            # 寫入 .env 檔
            env_lines = []
            if os.path.exists(env_path):
                with open(env_path, 'r', encoding='utf-8') as f:
                    env_lines = f.readlines()
            # 更新或新增
            new_lines = []
            found_secret = False
            found_token = False
            for line in env_lines:
                if line.startswith('LINE_CHANNEL_SECRET='):
                    new_lines.append(f'LINE_CHANNEL_SECRET={secret}\n')
                    found_secret = True
                elif line.startswith('LINE_CHANNEL_ACCESS_TOKEN='):
                    new_lines.append(f'LINE_CHANNEL_ACCESS_TOKEN={token}\n')
                    found_token = True
                else:
                    new_lines.append(line)
            if not found_secret:
                new_lines.append(f'LINE_CHANNEL_SECRET={secret}\n')
            if not found_token:
                new_lines.append(f'LINE_CHANNEL_ACCESS_TOKEN={token}\n')
            with open(env_path, 'w', encoding='utf-8') as f:
                f.writelines(new_lines)
            # 重新載入環境變數並初始化
            os.environ['LINE_CHANNEL_SECRET'] = secret
            os.environ['LINE_CHANNEL_ACCESS_TOKEN'] = token
            init_line(app)
            flash('LINE Bot 設定已更新，請重新啟動伺服器以完整生效', 'success')
        else:
            flash('請填寫完整的 Channel Secret 和 Access Token', 'danger')
        return redirect(url_for('line_settings'))

    current_secret = os.environ.get('LINE_CHANNEL_SECRET', '')
    current_token = os.environ.get('LINE_CHANNEL_ACCESS_TOKEN', '')
    configured = is_configured()
    sub_count = 0
    if configured:
        real = get_real_follower_count(app)
        sub_count = real if real is not None else get_subscriber_count(app)
    return render_template('line_settings.html',
                           configured=configured,
                           current_secret=current_secret,
                           current_token=current_token,
                           subscriber_count=sub_count)


# ----------------------------------------------------------
# 標的分析（admin only）
# ----------------------------------------------------------

@app.route('/analysis', methods=['GET', 'POST'])
@login_required
@admin_required
def analysis():
    results = None
    basket = None

    if request.method == 'POST':
        try:
            from fcn_quant_engine import FCNQuantEngine

            tickers_raw = request.form.get('tickers', '').strip().upper()
            tickers = [t.strip() for t in tickers_raw.replace('，', ',').split(',') if t.strip()]
            tenor = int(request.form.get('tenor', 6))
            ki_pct = int(request.form.get('ki_pct', 70)) / 100.0
            coupon = float(request.form.get('coupon', 12)) / 100.0
            investment = float(request.form.get('investment', '100,000').replace(',', ''))

            if not tickers:
                flash('請輸入至少一個標的代碼', 'warning')
                return render_template('analysis.html', results=None, basket=None)

            engine = FCNQuantEngine()

            if len(tickers) == 1:
                r = engine.analyze(tickers[0], tenor_months=tenor, ki_pct=ki_pct, coupon_rate=coupon)
                results = [r]
            else:
                basket = engine.analyze_basket(tickers, tenor_months=tenor, ki_pct=ki_pct)
                results = list(basket['individual'].values())
                # 補上 coupon 計算到每個結果
                for r in results:
                    r.coupon_rate = coupon
                    r.total_coupon_pct = round(coupon / 12 * tenor * 100, 2)
                    r.effective_cost_pct = round(r.ki_pct * (1 - coupon / 12 * tenor) * 100, 2)
                    r.effective_cost_price = round(r.current_price * r.effective_cost_pct / 100, 2)
                    r.breakeven_vs_current = round((1 - r.effective_cost_pct / 100) * 100, 2)

            # --- 歷史情境模擬 ---
            try:
                import yfinance as yf
                scenario_dates = [
                    ('2020-02-19', 'COVID 崩盤前（2020/02）'),
                    ('2022-01-03', '升息熊市前（2022/01）'),
                    ('2024-07-10', 'AI 泡沫疑慮（2024/07）'),
                ]
                tenor_days = tenor * 30
                for r in results:
                    scenarios = []
                    try:
                        stock = yf.Ticker(r.ticker)
                        for date_str, label in scenario_dates:
                            try:
                                start_dt = datetime.strptime(date_str, '%Y-%m-%d')
                                end_dt = start_dt + timedelta(days=tenor_days)
                                hist = stock.history(start=date_str, end=end_dt.strftime('%Y-%m-%d'))
                                if hist.empty or len(hist) < 5:
                                    continue
                                initial_price = round(float(hist['Close'].iloc[0]), 2)
                                min_price = round(float(hist['Low'].min()), 2)
                                max_drop_pct = round((min_price - initial_price) / initial_price * 100, 1)
                                ki_level = initial_price * ki_pct  # 執行價
                                hit_ki = min_price <= ki_level
                                scenario = {
                                    'label': label,
                                    'initial_price': initial_price,
                                    'min_price': min_price,
                                    'max_drop_pct': max_drop_pct,
                                    'ki_level': round(ki_level, 2),
                                    'hit_ki': hit_ki,
                                }
                                if hit_ki:
                                    # 接股價 = KI 價
                                    pickup_price = round(ki_level, 2)
                                    total_coupon_frac = coupon / 12 * tenor
                                    effective_cost = round(pickup_price * (1 - total_coupon_frac), 2)
                                    # 目前價格（最後一筆收盤價）
                                    latest = stock.history(period='1d')
                                    current_now = round(float(latest['Close'].iloc[-1]), 2) if not latest.empty else 0
                                    recovery_pct = round((current_now - effective_cost) / effective_cost * 100, 1) if effective_cost > 0 else 0
                                    scenario['pickup_price'] = pickup_price
                                    scenario['effective_cost'] = effective_cost
                                    scenario['current_now'] = current_now
                                    scenario['recovery_pct'] = recovery_pct
                                scenarios.append(scenario)
                            except Exception:
                                continue
                    except Exception:
                        pass
                    r.scenarios = scenarios
            except Exception:
                for r in results:
                    r.scenarios = []

            # --- 信心分級（基於 2026-04-14 50 檔回測 Brier）---
            try:
                from ticker_confidence import get_confidence, worst_tier
                for r in results:
                    tier, label, note, brier = get_confidence(r.ticker)
                    r.confidence_tier = tier
                    r.confidence_label = label
                    r.confidence_note = note
                    r.confidence_brier = brier
                if basket is not None:
                    wt, wt_list = worst_tier(tickers)
                    basket['confidence_tier'] = wt
                    basket['confidence_worst_tickers'] = wt_list
            except Exception:
                for r in results:
                    r.confidence_tier = 'unknown'
                    r.confidence_label = '未知'
                    r.confidence_note = ''
                    r.confidence_brier = None

            log_activity(f'標的分析：{",".join(tickers)} {tenor}M KI{int(ki_pct*100)}%')

        except Exception as e:
            flash(f'分析錯誤：{str(e)}', 'danger')

    return render_template('analysis.html', results=results, basket=basket)


# ----------------------------------------------------------
# 對沖成本計算器（admin only）
# ----------------------------------------------------------

@app.route('/hedge', methods=['GET', 'POST'])
@login_required
@admin_required
def hedge():
    result = None

    if request.method == 'POST':
        try:
            import yfinance as yf
            from datetime import datetime, timedelta

            ticker_raw = request.form.get('ticker', '').strip().upper()
            # 支援多檔（逗號分隔）
            tickers = [t.strip() for t in ticker_raw.replace('，', ',').split(',') if t.strip()]
            tenor_months = int(request.form.get('tenor', 6))
            ki_pct_int = int(request.form.get('ki_pct', 70))
            ki_pct = ki_pct_int / 100.0
            coupon_rate = float(request.form.get('coupon', 12))
            investment = float(request.form.get('investment', '100,000').replace(',', ''))

            if not tickers:
                flash('請輸入標的代碼', 'warning')
                return render_template('hedge.html', result=None)

            rf_rate = 4.5  # 無風險利率（年化）
            rf_tenor = round(rf_rate / 12 * tenor_months, 2)
            total_coupon_pct = round(coupon_rate / 12 * tenor_months, 2)
            coupon_income = round(investment * total_coupon_pct / 100, 0)

            # 每檔標的個別計算風險成本
            per_stock = []
            failed = []

            for tk in tickers:
                try:
                    stock = yf.Ticker(tk)
                    info = stock.info or {}
                    cp = float(info.get('currentPrice', 0) or info.get('regularMarketPrice', 0))
                    if cp <= 0:
                        failed.append(f'{tk}（無股價）')
                        continue

                    cn = info.get('shortName', tk)
                    kp = round(cp * ki_pct, 2)

                    exp_dates = stock.options
                    if not exp_dates:
                        failed.append(f'{tk}（無選擇權資料）')
                        continue

                    target_date = datetime.now() + timedelta(days=tenor_months * 30)
                    best_exp = min(exp_dates, key=lambda d: abs((datetime.strptime(d, '%Y-%m-%d') - target_date).days))

                    chain = stock.option_chain(best_exp)
                    puts = chain.puts
                    puts_valid = puts[puts['lastPrice'] > 0].copy()

                    if puts_valid.empty:
                        failed.append(f'{tk}（選擇權無報價）')
                        continue

                    # 找最接近 KI 價的 Put
                    ki_puts = puts_valid.iloc[(puts_valid['strike'] - kp).abs().argsort()[:1]]
                    if ki_puts.empty:
                        failed.append(f'{tk}（找不到對應 strike）')
                        continue

                    put_price = float(ki_puts.iloc[0]['lastPrice'])
                    put_strike = float(ki_puts.iloc[0]['strike'])
                    cost_pct = round(put_price / cp * 100, 2)

                    per_stock.append({
                        'ticker': tk,
                        'company_name': cn,
                        'current_price': cp,
                        'ki_price': kp,
                        'put_strike': put_strike,
                        'put_price': put_price,
                        'cost_pct': cost_pct,
                        'option_expiry': best_exp,
                    })
                except Exception as e:
                    failed.append(f'{tk}（錯誤：{str(e)[:30]}）')

            if not per_stock:
                result = {
                    'ticker': ticker_raw,
                    'company_name': '',
                    'is_basket': len(tickers) > 1,
                    'tickers': tickers,
                    'current_price': 0, 'ki_pct': ki_pct_int, 'ki_price': 0,
                    'tenor_months': tenor_months,
                    'coupon_rate': coupon_rate, 'total_coupon_pct': total_coupon_pct,
                    'investment': investment, 'coupon_income': coupon_income,
                    'option_expiry': '-', 'strategies': None,
                    'error_msg': f'無法取得風險成本：{", ".join(failed)}',
                    'conclusion': '', 'client_talk': '',
                }
                return render_template('hedge.html', result=result)

            is_basket = len(per_stock) > 1

            # === 單檔模式 ===
            if not is_basket:
                ps = per_stock[0]
                cp = ps['current_price']
                kp = ps['ki_price']
                put_strike = ps['put_strike']
                put_price = ps['put_price']
                shares = investment / cp

                strategies = []

                # 完全消除風險
                cost_pct = round(put_price / cp * 100, 2)
                cost_dollar = round(shares * put_price, 0)
                strategies.append({
                    'name': f'完全消除風險的市場成本（Put ${put_strike:.0f}）',
                    'description': f'消除股價跌破 ${put_strike:.0f} 的全部風險，市場報價 ${put_price:.2f}/股',
                    'cost_pct': cost_pct,
                    'cost_dollar': cost_dollar,
                    'net_yield': round(total_coupon_pct - cost_pct, 2),
                    'net_income': round(coupon_income - cost_dollar, 0),
                    'recommended': True,
                })

                # Put Spread
                try:
                    stock = yf.Ticker(ps['ticker'])
                    chain = stock.option_chain(ps['option_expiry'])
                    puts_valid = chain.puts[chain.puts['lastPrice'] > 0].copy()
                    deeper_target = cp * (ki_pct - 0.10)
                    deeper_puts = puts_valid.iloc[(puts_valid['strike'] - deeper_target).abs().argsort()[:1]]
                    if not deeper_puts.empty:
                        sell_price = float(deeper_puts.iloc[0]['lastPrice'])
                        sell_strike = float(deeper_puts.iloc[0]['strike'])
                        net_cost = max(put_price - sell_price, 0.01)
                        cost_pct = round(net_cost / cp * 100, 2)
                        cost_dollar = round(shares * net_cost, 0)
                        strategies.append({
                            'name': f'部分保護的市場成本（${put_strike:.0f} / ${sell_strike:.0f}）',
                            'description': f'保護 ${put_strike:.0f} 到 ${sell_strike:.0f} 區間的風險',
                            'cost_pct': cost_pct,
                            'cost_dollar': cost_dollar,
                            'net_yield': round(total_coupon_pct - cost_pct, 2),
                            'net_income': round(coupon_income - cost_dollar, 0),
                            'recommended': False,
                        })

                    disaster_target = cp * 0.55
                    disaster_puts = puts_valid.iloc[(puts_valid['strike'] - disaster_target).abs().argsort()[:1]]
                    if not disaster_puts.empty:
                        dis_price = float(disaster_puts.iloc[0]['lastPrice'])
                        dis_strike = float(disaster_puts.iloc[0]['strike'])
                        cost_pct = round(dis_price / cp * 100, 2)
                        cost_dollar = round(shares * dis_price, 0)
                        strategies.append({
                            'name': f'防大崩盤的市場成本（Put ${dis_strike:.0f}）',
                            'description': f'只衡量股價跌破 ${dis_strike:.0f}（腰斬級）的極端風險',
                            'cost_pct': cost_pct,
                            'cost_dollar': cost_dollar,
                            'net_yield': round(total_coupon_pct - cost_pct, 2),
                            'net_income': round(coupon_income - cost_dollar, 0),
                            'recommended': False,
                        })
                except Exception:
                    pass

                risk_premium_pct = strategies[0]['cost_pct']
                risk_premium_annual = round(risk_premium_pct / tenor_months * 12, 1)
                alpha_tenor = round(total_coupon_pct - rf_tenor - risk_premium_pct, 2)
                alpha_annual = round(alpha_tenor / tenor_months * 12, 1)
                rp_ratio = round(risk_premium_pct / total_coupon_pct * 100, 0) if total_coupon_pct > 0 else 0

                conclusion = (f"這檔 FCN 的風險溢價約 {risk_premium_pct:.2f}%（年化約 {risk_premium_annual:.1f}%）。"
                              f"票息 {coupon_rate:.1f}% 中，扣除無風險利率 {rf_rate}% 和風險溢價 {risk_premium_annual:.1f}%，"
                              f"客戶實際多賺的 Alpha 約 {alpha_annual:.1f}%。")
                if rp_ratio > 50:
                    conclusion += f" 風險溢價偏高（佔票息 {rp_ratio:.0f}%），建議審慎評估。"
                elif alpha_annual > 0:
                    conclusion += f" 風險溢價合理（佔票息 {rp_ratio:.0f}%），承擔風險後仍有正向 Alpha。"
                else:
                    conclusion += f" Alpha 為負，票息不足以補償風險，建議考慮其他條件。"

                client_talk = (f"這檔商品年化配息 {coupon_rate:.1f}%，比定存高出不少。"
                               f"其中約 {risk_premium_annual:.1f}% 是因為您承擔了股價跌破 {ki_pct_int}% 的風險。"
                               f"扣掉這部分，您還是比無風險利率多賺約 {max(alpha_annual, 0):.1f}%。")

                result = {
                    'ticker': ps['ticker'],
                    'company_name': ps['company_name'],
                    'is_basket': False,
                    'tickers': tickers,
                    'per_stock': per_stock,
                    'current_price': cp,
                    'ki_pct': ki_pct_int,
                    'ki_price': kp,
                    'tenor_months': tenor_months,
                    'coupon_rate': coupon_rate,
                    'total_coupon_pct': total_coupon_pct,
                    'investment': investment,
                    'coupon_income': coupon_income,
                    'option_expiry': ps['option_expiry'],
                    'strategies': strategies,
                    'conclusion': conclusion,
                    'client_talk': client_talk,
                    'error_msg': '',
                }

            # === 籃子模式（Worst-of）===
            else:
                # Worst-of 風險 ≈ 各股 Put 成本加總 × 相關性折扣
                # 實務上因為任一檔觸及就算 KI，風險通常高於單一股
                # 保守估計：總和 × 0.85（假設有一定相關性分散）
                total_cost_pct = sum(p['cost_pct'] for p in per_stock)
                basket_risk_pct = round(total_cost_pct * 0.85, 2)  # Worst-of 折扣

                # 找風險最高的股
                worst_stock = max(per_stock, key=lambda x: x['cost_pct'])

                shares_equiv = investment / worst_stock['current_price']

                strategies = []
                strategies.append({
                    'name': f'完全消除籃子風險的市場成本（{len(per_stock)} 檔 Worst-of）',
                    'description': (f'Worst-of 結構：任一檔跌破執行價即觸發。'
                                    f'消除全部風險需對每一檔都買 Put，加總 {total_cost_pct:.2f}%，'
                                    f'考慮相關性折扣後約 {basket_risk_pct:.2f}%'),
                    'cost_pct': basket_risk_pct,
                    'cost_dollar': round(investment * basket_risk_pct / 100, 0),
                    'net_yield': round(total_coupon_pct - basket_risk_pct, 2),
                    'net_income': round(coupon_income - investment * basket_risk_pct / 100, 0),
                    'recommended': True,
                })

                # 只保護最危險的那檔
                worst_cost = worst_stock['cost_pct']
                strategies.append({
                    'name': f'只保護最危險的 {worst_stock["ticker"]}（Put ${worst_stock["put_strike"]:.0f}）',
                    'description': f'{worst_stock["ticker"]} 風險成本最高（{worst_cost:.2f}%），先消除這個',
                    'cost_pct': worst_cost,
                    'cost_dollar': round(investment * worst_cost / 100, 0),
                    'net_yield': round(total_coupon_pct - worst_cost, 2),
                    'net_income': round(coupon_income - investment * worst_cost / 100, 0),
                    'recommended': False,
                })

                # Alpha 計算
                risk_premium_pct = basket_risk_pct
                risk_premium_annual = round(risk_premium_pct / tenor_months * 12, 1)
                alpha_tenor = round(total_coupon_pct - rf_tenor - risk_premium_pct, 2)
                alpha_annual = round(alpha_tenor / tenor_months * 12, 1)
                rp_ratio = round(risk_premium_pct / total_coupon_pct * 100, 0) if total_coupon_pct > 0 else 0

                conclusion = (f"這檔 Worst-of 籃子 FCN（{len(per_stock)} 檔）的風險溢價約 {risk_premium_pct:.2f}%"
                              f"（年化約 {risk_premium_annual:.1f}%）。各檔個別風險加總 {total_cost_pct:.2f}%，"
                              f"考慮相關性折扣後約 {basket_risk_pct:.2f}%。")
                conclusion += (f" 風險最高的是 {worst_stock['ticker']}（{worst_cost:.2f}%）。"
                               f"票息 {coupon_rate:.1f}% 中，扣除無風險利率和風險溢價後 "
                               f"Alpha 約 {alpha_annual:.1f}%。")

                if rp_ratio > 70:
                    conclusion += f" 風險溢價極高（佔票息 {rp_ratio:.0f}%），建議減少標的數或提高票息。"
                elif rp_ratio > 50:
                    conclusion += f" 風險溢價偏高（佔票息 {rp_ratio:.0f}%），建議審慎評估。"
                elif alpha_annual > 0:
                    conclusion += f" 風險溢價合理（佔票息 {rp_ratio:.0f}%）。"
                else:
                    conclusion += f" Alpha 為負，票息不足以補償這個籃子的風險。"

                client_talk = (f"這組 {len(per_stock)} 檔的籃子商品，任一檔跌破 {ki_pct_int}% 都會觸發。"
                               f"因為要承擔多檔的風險，年化 {coupon_rate:.1f}% 的配息裡有約 "
                               f"{risk_premium_annual:.1f}% 是承擔風險的報酬。"
                               f"扣掉這部分和定存利率，您實際多賺約 {max(alpha_annual, 0):.1f}%。"
                               f"其中最大的風險來自 {worst_stock['ticker']}。")

                # 基本資訊取第一檔的到期日
                result = {
                    'ticker': ticker_raw,
                    'company_name': f'{len(per_stock)} 檔 Worst-of 籃子',
                    'is_basket': True,
                    'tickers': tickers,
                    'per_stock': per_stock,
                    'worst_stock': worst_stock['ticker'],
                    'current_price': 0,  # 籃子沒有單一現價
                    'ki_pct': ki_pct_int,
                    'ki_price': 0,
                    'tenor_months': tenor_months,
                    'coupon_rate': coupon_rate,
                    'total_coupon_pct': total_coupon_pct,
                    'investment': investment,
                    'coupon_income': coupon_income,
                    'option_expiry': per_stock[0]['option_expiry'],
                    'strategies': strategies,
                    'conclusion': conclusion,
                    'client_talk': client_talk,
                    'error_msg': '',
                }

            log_activity(f'風險成本分析：{ticker_raw} {tenor_months}M KI{ki_pct_int}% 票息{coupon_rate}%')

        except Exception as e:
            flash(f'計算錯誤：{str(e)}', 'danger')

    return render_template('hedge.html', result=result)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', debug=(os.environ.get('FLASK_ENV') != 'production'), port=port)
