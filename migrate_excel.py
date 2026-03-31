"""
從 SN記憶表格.xlsx 匯入資料到 SQLite
"""
import sys, io, re, os
os.environ['PYTHONIOENCODING'] = 'utf-8'

from app import app, db
from models import Client, Product, Underlying, Position
import openpyxl
from datetime import datetime

XLSX = r'C:\Users\admin\Desktop\新增資料夾 (3)\SN記憶表格.xlsx'


def parse_clients_from_notes(notes_list):
    """從備注欄解析客戶：'朱慶瑋6萬' → [('朱慶瑋', 60000)]"""
    results = []
    for note in notes_list:
        if not note or not isinstance(note, str):
            continue
        note = note.strip()
        if not note:
            continue
        m = re.match(r'(.+?)(\d+)萬', note)
        if m:
            name = m.group(1)
            amount = int(m.group(2)) * 10000
            results.append((name, amount))
    return results


def get_or_create_client(name):
    c = Client.query.filter_by(name=name).first()
    if not c:
        c = Client(name=name, name_masked=Client.mask_name(name))
        db.session.add(c)
        db.session.flush()
    return c


def parse_sheet(ws, status):
    """解析一個 sheet，每 5 行一組商品"""
    max_row = ws.max_row
    r = 2  # 從第二行開始（第一行是表頭）
    count = 0

    while r <= max_row:
        code_cell = ws.cell(r, 2).value
        if not code_cell or not isinstance(code_cell, str) or 'SN' not in code_cell:
            r += 1
            continue

        product_code = code_cell.strip()

        # 檢查是否已存在
        if Product.query.filter_by(product_code=product_code).first():
            print(f'  跳過（已存在）: {product_code}')
            r += 5
            continue

        # Row 1: 商品基本資料
        tenor      = ws.cell(r, 4).value
        trade_date = ws.cell(r, 16).value
        start_date = ws.cell(r, 17).value
        mat_date   = ws.cell(r, 18).value
        rate       = ws.cell(r, 19).value
        col_e_r1   = ws.cell(r, 5).value   # 交易營業員/備注

        # 標的 ticker（Row 1 的 I, K, M, O 欄）
        tickers = [
            ws.cell(r, 9).value,   # Underlying 1
            ws.cell(r, 11).value,  # Underlying 2
            ws.cell(r, 13).value,  # Underlying 3
            ws.cell(r, 15).value,  # Underlying 4
        ]

        # Row 2: 期初價格
        prices = [
            ws.cell(r+1, 9).value,
            ws.cell(r+1, 11).value,
            ws.cell(r+1, 13).value,
            ws.cell(r+1, 15).value,
        ]

        # Row 3: KO
        ko_pct = ws.cell(r+2, 7).value
        ko_hits = [
            ws.cell(r+2, 8).value,   # H
            ws.cell(r+2, 10).value,  # J
            ws.cell(r+2, 12).value,  # L
            ws.cell(r+2, 14).value,  # N
        ]

        # Row 4: Strike
        strike_pct = ws.cell(r+3, 7).value

        # Row 5: EKI
        eki_label = ws.cell(r+4, 6).value
        eki_pct   = ws.cell(r+4, 7).value

        # 判斷是否有 EKI
        has_eki = True
        if eki_label and '無' in str(eki_label):
            has_eki = False
            eki_pct = None
        if eki_pct == 0:
            has_eki = False
            eki_pct = None

        # 判斷 KO type（從備注欄）
        ko_type = 'fixed'
        special_notes = ''
        if col_e_r1 and isinstance(col_e_r1, str):
            if 'stepdown' in col_e_r1.lower():
                ko_type = 'stepdown'
                special_notes = col_e_r1.strip()
            elif col_e_r1.strip() in ('AUD', 'EUR', 'TWD'):
                special_notes = col_e_r1.strip()

        # 收集比價日備注（stepdown 產品的觀察日在 E 欄）
        for dr in range(1, 5):
            cell_e = ws.cell(r+dr, 5).value
            if cell_e and isinstance(cell_e, str) and '/' in cell_e:
                if special_notes:
                    special_notes += '\n'
                special_notes += cell_e.strip()

        # 判斷幣別
        currency = 'USD'
        if special_notes and 'AUD' in special_notes.upper():
            currency = 'AUD'

        # 日期轉換
        def to_date(v):
            if isinstance(v, datetime):
                return v.date()
            return None

        p = Product(
            product_code  = product_code,
            product_type  = 'FCN',
            tenor_months  = int(tenor) if tenor else None,
            currency      = currency,
            coupon_rate   = float(rate) if rate else None,
            trade_date    = to_date(trade_date),
            start_date    = to_date(start_date),
            maturity_date = to_date(mat_date),
            ko_pct        = float(ko_pct) if ko_pct else None,
            strike_pct    = float(strike_pct) if strike_pct else None,
            eki_pct       = float(eki_pct) if eki_pct and has_eki else None,
            ko_type       = ko_type,
            special_notes = special_notes.strip() if special_notes else None,
            status         = status,
        )
        db.session.add(p)
        db.session.flush()

        # 建立標的
        for i in range(4):
            ticker = tickers[i]
            if not ticker or not isinstance(ticker, str):
                continue
            init_p = float(prices[i]) if prices[i] and isinstance(prices[i], (int, float)) else None
            ko_lv  = round(init_p * p.ko_pct, 4) if init_p and p.ko_pct else None
            st_lv  = round(init_p * p.strike_pct, 4) if init_p and p.strike_pct else None
            eki_lv = round(init_p * p.eki_pct, 4) if init_p and p.eki_pct else None

            hit = str(ko_hits[i]).upper().strip() == 'V' if ko_hits[i] else False

            u = Underlying(
                product_id     = p.id,
                ticker         = ticker.strip().upper(),
                initial_price  = init_p,
                ko_level       = ko_lv,
                strike_level   = st_lv,
                eki_level      = eki_lv,
                ko_hit         = hit,
                position_order = i + 1,
            )
            db.session.add(u)

        # 解析客戶部位（備注欄散落在各行各處）
        client_notes = []
        for dr in range(5):
            val_t = ws.cell(r + dr, 20).value  # 第 T 欄
            if val_t and isinstance(val_t, str) and '萬' in val_t:
                client_notes.append(val_t)
        # Row 1 的 T 欄
        val_t1 = ws.cell(r, 20).value
        if val_t1 and isinstance(val_t1, str) and '萬' in val_t1 and val_t1 not in client_notes:
            client_notes.append(val_t1)

        clients_parsed = parse_clients_from_notes(client_notes)
        for name, amount in clients_parsed:
            client = get_or_create_client(name)
            pos = Position(
                client_id         = client.id,
                product_id        = p.id,
                investment_amount = amount,
            )
            db.session.add(pos)

        count += 1
        print(f'  匯入: {product_code} ({len([t for t in tickers if t])} 標的, {len(clients_parsed)} 客戶)')
        r += 5

    return count


with app.app_context():
    db.create_all()

    wb = openpyxl.load_workbook(XLSX)

    # 則興 sheet → active
    print('=== 匯入 則興 (active) ===')
    c1 = parse_sheet(wb['則興'], 'active')

    # KO sheet → ko_exited
    print('\n=== 匯入 KO (ko_exited) ===')
    c2 = parse_sheet(wb['KO'], 'ko_exited')

    db.session.commit()
    print(f'\n完成！共匯入 {c1 + c2} 筆商品')
    print(f'客戶數：{Client.query.count()}')
    print(f'商品數：{Product.query.count()}')
    print(f'部位數：{Position.query.count()}')
